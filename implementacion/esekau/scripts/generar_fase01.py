#!/usr/bin/env python3
"""Fase 1 demo esekau: universo + recorte 80–100 SKUs descriptivos.

schema_name = esekau. Precio Final imp. incl. = lista 1 (reventa).
UMV = unidad (el precio es por unidad; caja/bulto va en aliases).
"""
from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INP = ROOT / "inputs" / "Esekau cordoba lista productos Suplai.xlsx"
OUT = ROOT / "outputs"
FULL = ROOT / "inputs" / "catalogo-completo.csv"

TARGET_MIN = 80
TARGET_MAX = 100

# Recorte demo ~95. Orden = rotación (P&G primero).
DEMO_CODES = [
    # P&G / Pantene (líder) + HS + Oral-B + Pampers + Always + Downy + Gillette
    "1079",
    "15006",
    "116",
    "118",
    "1453",
    "1451",
    "17500",
    "1871",
    "552",
    "551",
    "7636",
    "1880",
    "8006",
    "8105",
    "8106",
    "1194",
    "1191",
    "1192",
    "6320",
    "1088",
    "11797",
    "2413",
    "1095",
    "1098",
    "11234",
    "880",
    "550",
    "1430",
    "220",
    "2248",
    "11617",
    "1003",
    # Dreamco
    "1641",
    "1642",
    "1643",
    "18729",
    "18730",
    "18359",
    "18353",
    "18342",
    "18343",
    "18387",
    "18389",
    "18240",
    "18260",
    "18416",
    "18415",
    "2082",
    "18367",
    "18707",
    "18400",
    # Clorox
    "20102",
    "20104",
    "20109",
    "20032",
    "20035",
    "20410",
    "20305",
    "20193",
    # Cepas
    "10108",
    "10109",
    "10107",
    "10112",
    "10114",
    "10105",
    "10043",
    "10040",
    "10044",
    "10050",
    "10065",
    "10077",
    "10069",
    "10053",
    "10054",
    "10003",
    "10017",
    "10240",
    "10056",
    "508",
    "10132",
    # Varios / almacén
    "11011",
    "11012",
    "11071",
    "18805",
    "18928",
    "18828",
    "11300",
    "1014",
    "1016",
    "1015",
    "11279",
    "11282",
    "4003",
    "11685",
    "5603",
]

PRODUCT_FIELDS = [
    "product_code",
    "nombre",
    "precio_lista_1",
    "stock",
    "unidades_por_bulto",
    "unidad_minima_de_venta",
    "umv_tipo",
    "categoria_1",
    "categoria_2",
    "categoria_3",
    "categoria_4",
    "aliases",
    "rotacion_index",
    "mental_priority",
    "descripcion",
    "image_url",
    "en_catalogo",
    "is_mock",
    "fuente_hoja",
]

PRICE_FIELDS = [
    "lista_precios_id",
    "nombre",
    "multiplicador_sobre_lista_1",
    "product_code",
    "precio_unidad",
    "is_mock",
]

LISTAS = [
    (1, "Lista Base Córdoba", 1.00),
    (2, "Lista Minorista Sugerido", 1.15),
    (3, "Lista Mayorista Especial", 0.90),
    (4, "Lista Gran Distribuidor", 0.85),
]

MARCAS = [
    ("HEAD & SHOULDERS", "Head & Shoulders"),
    ("H&S", "Head & Shoulders"),
    ("PANTENE", "Pantene"),
    ("ORAL-B", "Oral-B"),
    ("ORAL B", "Oral-B"),
    ("PAMPERS", "Pampers"),
    ("ALWAYS", "Always"),
    ("DOWNY", "Downy"),
    ("PRESTOBARBA", "Prestobarba"),
    ("PRESTO ", "Prestobarba"),
    ("PRESTO3", "Prestobarba"),
    ("GILLETTE", "Gillette"),
    ("VENUS", "Venus"),
    ("OLD SPICE", "Old Spice"),
    ("SECRET", "Secret"),
    ("FOAMY", "Foamy"),
    ("MACH3", "Gillette Mach3"),
    ("MAGISTRAL", "Magistral"),
    ("ZORRO", "Zorro"),
    ("ARIEL", "Ariel"),
    ("ACE ", "Ace"),
    ("ACE POUCH", "Ace"),
    ("ACE BOTELLA", "Ace"),
    ("ACE CLASI", "Ace"),
    ("ACE DILUIBLE", "Ace"),
    ("PLB ", "Plusbelle"),
    ("FEDERAL", "Federal"),
    ("LIMZUL", "Limzul"),
    ("POETT", "Poett"),
    ("AYUDIN", "Ayudín"),
    ("TRENET", "Trenet"),
    ("SELTON", "Selton"),
    ("MORTIMER", "Mortimer"),
    ("TERMA", "Terma"),
    ("DR LEMON", "Dr Lemon"),
    ("BACARDI", "Bacardi"),
    ("JAMAICA", "Jamaica"),
    ("NIKOV", "Nikov"),
    ("BOMBAY", "Bombay"),
    ("AMARULA", "Amarula"),
    ("FERNET 1882", "Fernet 1882"),
    ("GANCIA", "Gancia"),
    ("MARTINI", "Martini"),
    ("MI CHANCES", "Mi Chances"),
    ("COCA COLA", "Coca-Cola"),
    ("SPRITE", "Sprite"),
    ("TARAGUI", "Taragüi"),
    ("UNION", "Unión"),
    ("MAÑANITA", "Mañanita"),
    ("CHOCMAN", "Chocman"),
    ("NUGATON", "Nugaton"),
    ("PRINGLES", "Pringles"),
    ("DURACELL", "Duracell"),
    ("PHILCO", "Philco"),
    ("SUEROX", "Suerox"),
    ("PRIME", "Prime"),
    ("PATITO", "Patito"),
    ("FELFORT", "Felfort"),
    ("OKEBON", "Okebon"),
    ("TRES TORRES", "Tres Torres"),
    ("TACCONI", "Tacconi"),
]


def norm_code(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def parse_qty(v) -> int:
    if v is None:
        return 1
    s = str(v).strip().replace(",", ".")
    try:
        n = int(float(s))
        return n if n > 0 else 1
    except ValueError:
        return 1


def marca_de(nombre: str) -> str:
    u = nombre.upper()
    if re.match(r"^HS\b", u) or u.startswith("HS "):
        return "Head & Shoulders"
    if u.startswith("CEP."):
        return "Oral-B"
    if u.startswith("PASTA ORAL"):
        return "Oral-B"
    if u.startswith("YM TARAGUI"):
        return "Taragüi"
    if u.startswith("YM UNION"):
        return "Unión"
    if u.startswith("YM MAÑANITA") or u.startswith("YM MANANITA"):
        return "Mañanita"
    if u.startswith("LAVANDINA"):
        return "Ayudín"
    if "ACE " in u or u.startswith("ACE"):
        return "Ace"
    for needle, label in MARCAS:
        if needle in u:
            return label
    return ""


def cat1_de(linea: str) -> str:
    m = {
        "PROCTER": "Cuidado personal",
        "DREAMCO": "Limpieza hogar",
        "CLOROX": "Limpieza hogar",
        "CEPAS": "Bebidas",
        "VARIOS": "Almacén",
    }
    return m.get((linea or "").upper(), linea or "General")


def cat2_de(rubro: str) -> str:
    r = (rubro or "").strip()
    r_compact = re.sub(r"\s+", " ", r)
    if "TOALLAS FEMENINAS" in r.upper() and "PAÑAL" in r.upper():
        return "Toallitas bebé"
    r = r_compact.replace("CUIDADADO", "CUIDADO")
    mapa = {
        "CUIDADO DE CABELLO (HAIR CARE )": "Cuidado de cabello",
        "CUIDADO BUCAL ( ORAL CARE )": "Cuidado bucal",
        "CUIDADO FEMENINO ( FEM CARE )": "Cuidado femenino",
        "DESECHABLES (GROOMING)": "Afeitado desechable",
        "SISTEMAS (GROOMING)": "Afeitado sistemas",
        "PREPARADOS DE AFEITAR ( GROOMING )": "Espuma de afeitar",
        "PAÑALES ( BABY CARE )": "Pañales",
        "PAÑALES  -  TOALLAS FEMENINAS": "Toallitas bebé",
        "PAÑALES - TOALLAS FEMENINAS": "Toallitas bebé",
        "SUAVIZANTE ( FABRIC ENHANCERS)": "Suavizante",
        "DESODORANTES ( APDOS )": "Desodorantes",
        "PILAS Y BATERIAS": "Pilas",
        "PILAS ( BATTERIES )": "Pilas",
        "SHAMPOO Y ACONDICIONADOR": "Shampoo y acondicionador",
        "LAVAVAJILLAS": "Lavavajillas",
        "JAB LIQUIDO": "Jabón líquido",
        "JAB COMPACTO": "Jabón en pan",
        "JAB TOCADOR INDIVIDUAL": "Jabón tocador",
        "SUAVIZANTES DR": "Suavizante",
        "LIMPIADORES LIQUIDOS": "Limpiadores",
        "LIMPIADORES ESPECIFICOS DR": "Limpiadores específicos",
        "DETERGENTES ( HOME CARE )": "Detergente",
        "POLVOS": "Detergente en polvo",
        "QUITAMANCHAS": "Quitamanchas",
        "PANIFICACION": "Galletitas",
        "TRATAMIENTOS Y CREMAS DE PEINAR": "Cremas de peinar",
        "CUIDADO DE LA ROPA": "Quitamanchas ropa",
        "LAVANDINAS": "Lavandina",
        "UTENSILIOS DE LIMPIEZA": "Utensilios",
        "REPELENTE": "Repelente",
        "ALTO GRADO": "Bebidas blancas",
        "LEMON": "Ready to drink",
        "TERMA": "Amargos",
        "VINOS": "Vinos",
        "FERNET": "Aperitivos",
        "GANCIA": "Aperitivos",
        "RESTO APERITIVOS": "Aperitivos",
        "RTD (PRONTO + 1882)": "Ready to drink",
        "AMARGOS VALUE": "Amargos",
        "OTROS": "Gaseosas",
        "YERBAS": "Yerba mate",
        "CHOCOLATE BONAFIDE": "Chocolate",
        "CHOCOLATE FELFORT": "Chocolate",
        "CHOCOLATE PASCUAS": "Chocolate",
        "BEBIDAS NO ALCOHOLICAS": "Bebidas isotónicas",
        "LATEX": "Preservativos",
        "LATEX M": "Preservativos",
        "TABACO": "Tabaco",
        "CAFE BONAFIDE": "Café",
        "CEREALES": "Cereales",
        "LIMPIEZA": "Guantes",
        "COSMETICA": "Cosmética",
    }
    return mapa.get(r, r.title() if r else "General")


def _has_sh(u: str) -> bool:
    return bool(re.search(r"(?:^|\s)SH(?:\s|$)", u) or "SHAMPOO" in u or u.startswith("PLB SH"))


def tipo_producto(nombre: str, cat2: str) -> str:
    u = nombre.upper()
    if "SACHET" in u:
        return "Sachet de shampoo" if _has_sh(u) else "Sachet"
    if "TOALLITA" in u:
        return "Toallitas húmedas"
    if "PAÑAL" in u or "PAMPERS BABY" in u or "BABYDRY" in u or "BABYSAN" in u:
        return "Pañal"
    if "CEP." in u or "CEPILLO" in u:
        return "Cepillo dental"
    if "PASTA" in u:
        return "Pasta dental"
    if "DES." in u or "ANT." in u or "ROLL ON" in u or "DEODOR" in u or "BARRA FRESH" in u:
        return "Desodorante"
    if " CN " in u or re.search(r"(?:^|\s)AC(?:\s|$)", u) or "ACONDICIONADOR" in u:
        return "Acondicionador"
    if _has_sh(u):
        return "Shampoo"
    if "CREM" in u and "PEIN" in u:
        return "Crema de peinar"
    if "SUAVIZ" in u or "DOWNY" in u:
        return "Suavizante de ropa"
    if ("GATILLO" in u or re.search(r"\bGAT\b", u)) and "BAÑO" in u:
        return "Limpiador de baño"
    if "GATILLO" in u or re.search(r"\bGAT\b", u):
        return "Limpiador en gatillo"
    if cat2 == "Lavavajillas" or "LAVAVAJ" in u or "ZORRO ULTRA DETERG" in u:
        return "Lavavajillas"
    if "MAGISTRAL" in u and cat2 != "Limpiadores":
        return "Lavavajillas"
    if "LAVANDINA" in u:
        return "Lavandina"
    if "POETT" in u:
        return "Limpiador desinfectante"
    if "JB LIQ" in u or "JABON EN" in u or "JAB." in u or "ARIEL" in u or re.search(r"\bACE\b", u):
        if "PAN" in u:
            return "Jabón en pan"
        if "POLVO" in u:
            return "Detergente en polvo"
        return "Jabón líquido para ropa"
    if "ALWAYS" in u and "DIARIO" in u:
        return "Protectores diarios"
    if "ALWAYS" in u:
        return "Toalla femenina"
    if "FOAMY" in u:
        return "Espuma de afeitar"
    if "VENUS" in u or "PRESTO" in u or ("GILLETTE" in u and "MAQ" in u):
        return "Máquina de afeitar"
    if "PILA" in u:
        return "Pila"
    if "TERMA" in u:
        return "Amargo herbal"
    if "DR LEMON" in u:
        return "Bebida lista para tomar"
    if "FERNET" in u:
        return "Fernet"
    if "GANCIA" in u or "MARTINI" in u:
        return "Aperitivo"
    if "BACARDI" in u or "JAMAICA" in u:
        return "Ron"
    if "VODKA" in u:
        return "Vodka"
    if re.search(r"\bGIN\b", u):
        return "Gin"
    if "MALBEC" in u or "CABERNET" in u or "VINO" in u or "CHANCES" in u:
        return "Vino"
    if "COCA" in u or "SPRITE" in u:
        return "Gaseosa"
    if "YM " in u or "YERBA" in u or "TARAGUI" in u:
        return "Yerba mate"
    if "CHOCMAN" in u or "NUGATON" in u or "FELFORT" in u or "CHOCOLATE" in u:
        return "Chocolate"
    if "PRINGLES" in u:
        return "Snack de papa"
    if "SUEROX" in u:
        return "Bebida isotónica"
    if "PRIME" in u or "MAXX" in u:
        return "Preservativo"
    if "GUANTE" in u:
        return "Guante de látex"
    if "ESPONJA" in u:
        return "Esponja"
    if "REPELENTE" in u:
        return "Repelente"
    if "OKEBON" in u or "PANAL" in u:
        return "Galletita"
    return cat2 or "Producto"


def formato_de(nombre: str) -> str:
    u = nombre.upper().replace(",", ".")
    if re.search(r"1\s*/\s*2\s*KG", u):
        return "1/2 kg"
    if re.search(r"1\.5\s*CC", u) or re.search(r"X\s*1\.5\s*CC", u):
        return "1.5 L"
    m = re.search(r"(\d+(?:\.\d+)?)\s*(ML|LTR|L\b|GRS?|G\b|KG|CC)\b", u)
    if m:
        n, unit = m.group(1), m.group(2)
        unit_map = {"ML": "ml", "LTR": "L", "L": "L", "GRS": "g", "GR": "g", "G": "g", "KG": "kg", "CC": "cc"}
        return f"{n} {unit_map.get(unit, unit.lower())}"
    m = re.search(r"X\s*(\d+(?:\.\d+)?)\s*(KG|GRS?|G|ML|L)\b", u)
    if m:
        return f"{m.group(1)} {m.group(2).lower()}"
    m = re.search(r"(\d+)\s*(UNI|UNID|UN)\b", u)
    if m:
        return f"{m.group(1)} un"
    return ""


def categorias(linea: str, rubro: str, nombre: str) -> tuple[str, str, str, str]:
    c1 = cat1_de(linea)
    c2 = cat2_de(rubro)
    c3 = marca_de(nombre) or linea.title()
    c4 = formato_de(nombre) or tipo_producto(nombre, c2)
    return c1, c2, c3, c4


def word_count(text: str) -> int:
    return len(text.replace(".", " ").split())


def descripcion_corta(nombre: str, linea: str, rubro: str, qty: int) -> str:
    cat2 = cat2_de(rubro)
    tipo = tipo_producto(nombre, cat2)
    marca = marca_de(nombre)
    fmt = formato_de(nombre)
    bits = [tipo]
    if marca:
        bits.append(f"marca {marca}")
    if fmt:
        bits.append(fmt)
    bits.append(f"caja x{qty}" if qty > 1 else "venta por unidad")
    bits.append(cat2.lower())
    u = nombre.upper()
    for extra, label in (
        ("LISO INFINITO", "liso infinito"),
        ("KERATINA", "keratina"),
        ("RESTAURACION", "restauración"),
        ("ANTICOMEZON", "anticómezón"),
        ("MANZANA", "manzana"),
        ("BRISA SUAVE", "brisa suave"),
        ("BRISA INTENSO", "brisa intenso"),
        ("NOCHES", "noche"),
        ("DIARIOS", "uso diario"),
        ("LIMON", "limón"),
        ("NARANJA", "naranja"),
        ("CITRUS", "citrus"),
        ("SERRANO", "serrano"),
        ("POMELO", "pomelo"),
        ("CUYANO", "cuyano"),
        ("BEBE", "bebé"),
        ("LAVANDA", "lavanda"),
        ("GLACIAR", "glaciar"),
        ("BABYDRY PEQ", "talle P"),
        ("BABYDRY MED", "talle M"),
        ("BABYDRY GDE", "talle G"),
        ("TEXTURADO", "texturado"),
        ("COCO", "coco"),
        ("MENTITAS", "mentitas"),
        ("CARTA BLANCA", "carta blanca"),
        ("GOLD RUM", "gold"),
        ("CON COLA", "con cola"),
        ("ROSSO", "rosso"),
        ("FRESH", "fresh"),
    ):
        if extra in u and label not in " ".join(bits).lower():
            bits.append(label)
            break
    text = bits[0] + ", " + ", ".join(bits[1:]) + "."
    if word_count(text) < 10:
        text = text.rstrip(".") + ", lista ESEKA-U sucursal Córdoba."
    words = text.split()
    if len(words) > 25:
        text = " ".join(words[:24]).rstrip(",.") + "."
    return text[0].upper() + text[1:]


def aliases(nombre: str, code: str, marca: str, qty: int, cat2: str) -> str:
    parts = [
        nombre,
        nombre.lower(),
        code,
        cat2.lower(),
        "caja",
        "bulto",
        "unidad",
    ]
    if qty > 1:
        parts += [f"caja x{qty}", f"bulto x{qty}", f"x{qty}"]
    if marca:
        low = marca.lower()
        parts += [marca, low, f"caja {low}", f"bulto {low}"]
    u = nombre.upper()
    if "SACHET" in u:
        parts += ["sachet", "sobrecito"]
    if "POUCH" in u:
        parts += ["pouch", "doypack"]
    if any(x in u for x in ("SH ", "SHAMPOO")):
        parts.append("shampoo")
    if "LAVAVAJ" in u or "MAGISTRAL" in u:
        parts += ["detergente", "lavavajillas", "det"]
    if "LAVANDINA" in u:
        parts += ["cloro", "lavandina"]
    if "YM " in u:
        parts += ["yerba", "yerba mate", "medio kilo" if "1/2" in u or "0.5" in u else ""]
    if "1.35L" in u or "1.35 L" in u:
        parts.append("litro y medio")
    for tok in re.findall(r"[A-ZÁÉÍÓÚÑ0-9]{3,}", nombre):
        parts.append(tok.lower())
    seen: set[str] = set()
    out: list[str] = []
    for p in parts:
        p = (p or "").strip()
        key = p.lower()
        if not p or key in seen:
            continue
        seen.add(key)
        out.append(p)
    return "|".join(out[:14])


def rotacion_full(linea: str, nombre: str, i: int, n: int) -> float:
    base = 0.90 - (i / max(n - 1, 1)) * 0.75
    u = nombre.upper()
    if linea == "PROCTER" and any(x in u for x in ("PANTENE", "PAMPERS", "ALWAYS", "DOWNY", "ORAL")):
        base = max(base, 0.78)
    return round(min(0.95, max(0.12, base)), 2)


def to_row(item: dict, i: int, n: int) -> dict:
    nombre = item["nombre"]
    linea = item["linea"]
    rubro = item["rubro"]
    qty = item["qty"]
    code = item["code"]
    cat1, cat2, cat3, cat4 = categorias(linea, rubro, nombre)
    rot = rotacion_full(linea, nombre, i, n)
    stock = int(40 + rot * 360)
    marca = marca_de(nombre)
    return {
        "product_code": code,
        "nombre": nombre,
        "precio_lista_1": f"{item['precio']:.2f}",
        "stock": str(stock),
        "unidades_por_bulto": str(qty),
        "unidad_minima_de_venta": "unidad",
        "umv_tipo": "unidad",
        "categoria_1": cat1,
        "categoria_2": cat2,
        "categoria_3": cat3,
        "categoria_4": cat4,
        "aliases": aliases(nombre, code, marca, qty, cat2),
        "rotacion_index": str(rot),
        "mental_priority": str(round(rot * 0.88, 2)),
        "descripcion": descripcion_corta(nombre, linea, rubro, qty),
        "image_url": "",
        "en_catalogo": "true",
        "is_mock": "true",
        "fuente_hoja": "Sucursal Córdoba",
    }


def write_csv(path: Path, fields: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in fields})


def load_excel() -> list[dict]:
    wb = load_workbook(INP, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    linea = ""
    rubro = ""
    items: list[dict] = []
    seen: set[str] = set()
    for r in ws.iter_rows(values_only=True):
        desc = str(r[0] or "").strip()
        if desc.upper().startswith("LINEA:"):
            linea = str(r[1] or "").strip()
            continue
        if desc.upper().startswith("RUBRO:"):
            rubro = str(r[1] or "").strip()
            continue
        if desc in ("", "Descripción", "ESEKA-U", "Sucursal Cordoba"):
            continue
        code = norm_code(r[1])
        if not code:
            continue
        try:
            precio = float(r[4] or 0)
        except (TypeError, ValueError):
            precio = 0.0
        if precio <= 0:
            continue
        if code in seen:
            continue
        seen.add(code)
        nombre = re.sub(r"\s+", " ", desc).strip()
        nombre = re.sub(r"[\-–—\.]{2,}$", "", nombre).strip()
        items.append(
            {
                "code": code,
                "nombre": nombre,
                "precio": precio,
                "qty": parse_qty(r[3]),
                "barcode": str(r[2] or "").strip(),
                "linea": linea,
                "rubro": rubro,
            }
        )
    wb.close()
    return items


def main() -> int:
    priced = load_excel()
    print(f"[*] origen con Precio Final > 0: {len(priced)}")

    full_rows: list[dict] = []
    by_code: dict[str, dict] = {}
    for i, item in enumerate(priced):
        row = to_row(item, i, len(priced))
        full_rows.append(row)
        by_code[item["code"]] = item
    write_csv(FULL, PRODUCT_FIELDS, full_rows)
    print(f"[*] catalogo-completo.csv: {len(full_rows)}")

    missing = [c for c in DEMO_CODES if c not in by_code]
    if missing:
        print(f"[FAIL] códigos del recorte ausentes: {missing}")
        return 1
    dupes = [c for c in DEMO_CODES if DEMO_CODES.count(c) > 1]
    if dupes:
        print(f"[FAIL] códigos duplicados en recorte: {sorted(set(dupes))}")
        return 1

    n = len(DEMO_CODES)
    if not (TARGET_MIN <= n <= TARGET_MAX):
        print(f"[FAIL] recorte fuera de 80–100: {n}")
        return 1

    demo: list[dict] = []
    counts: dict[str, int] = defaultdict(int)
    marcas: dict[str, int] = defaultdict(int)
    for i, code in enumerate(DEMO_CODES):
        src = by_code[code]
        row = to_row(src, i, n)
        # Recorte: P&G alto, resto Pareto suave
        rot = round(0.95 - i * (0.40 / max(n - 1, 1)), 2)
        if src["linea"] == "PROCTER":
            rot = min(0.95, rot + 0.04)
        row["rotacion_index"] = f"{rot:.2f}"
        row["mental_priority"] = f"{round(rot * 0.88, 2):.2f}"
        row["stock"] = str(int(120 + rot * 280))
        demo.append(row)
        counts[src["linea"]] += 1
        marcas[marca_de(src["nombre"]) or src["linea"]] += 1

    write_csv(OUT / "phase-01-productos.csv", PRODUCT_FIELDS, demo)
    print(f"[*] phase-01-productos.csv: {len(demo)} SKUs demo")
    print("[*] líneas:", dict(counts))
    print("[*] marcas:", dict(sorted(marcas.items(), key=lambda x: -x[1])[:12]))

    short_desc = [p["product_code"] for p in demo if word_count(p["descripcion"]) < 10]
    long_desc = [p["product_code"] for p in demo if word_count(p["descripcion"]) > 25]
    if short_desc or long_desc:
        print(f"[WARN] descripciones cortas={short_desc} largas={long_desc}")

    combined_prices: list[dict] = []
    for list_id, nombre, mult in LISTAS:
        price_rows = []
        for p in demo:
            rec = {
                "lista_precios_id": list_id,
                "nombre": nombre,
                "multiplicador_sobre_lista_1": f"{mult:.2f}",
                "product_code": p["product_code"],
                "precio_unidad": f"{round(float(p['precio_lista_1']) * mult, 2):.2f}",
                "is_mock": "true",
            }
            price_rows.append(rec)
            combined_prices.append(rec)
        write_csv(OUT / f"phase-01-lista-precios-{list_id}.csv", PRICE_FIELDS, price_rows)
        print(f"[*] lista {list_id} {nombre}: {len(price_rows)}")
    write_csv(OUT / "phase-01-listas-precios.csv", PRICE_FIELDS, combined_prices)

    print("\n--- RECORTE DEMO ---")
    for p in demo:
        print(
            f"  {p['product_code']:>6}  ${float(p['precio_lista_1']):>10.2f}  "
            f"x{p['unidades_por_bulto']:<4} {p['nombre'][:58]}"
        )
        print(f"          {p['descripcion']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
