#!/usr/bin/env python3
"""Fase 1 demo Ramal: recorte 80–100 SKUs descriptivos + catalogo-completo.csv.

UMV: caja (= bulto interno). Precio Final Chess = precio de la caja.
Marca líder Marinaro (exclusiva) + mix de líneas del negocio.
"""
from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INP = ROOT / "inputs" / "plantilla-precios-completa.xlsx"
OUT = ROOT / "outputs"
FULL = ROOT / "inputs" / "catalogo-completo.csv"

TARGET_MIN = 80
TARGET_MAX = 100

# Cupos del recorte (orientativos; Marinaro entra casi entero = gancho de la demo).
QUOTAS = {
    "marinaro": 43,
    "sierra_padres": 4,
    "vinos_puna": 6,
    "vinos": 8,
    "gaseosas": 6,
    "cervezas": 8,
    "jugos": 5,
    "amargos": 4,
    "aguas": 3,
    "blancas": 5,
    "sidra": 2,
    "snacks": 3,
    "yerba": 2,
    "otros": 0,
}

PRODUCT_FIELDS = [
    "product_code",
    "nombre",
    "precio_lista_1",
    "stock",
    "unidades_por_bulto",
    "unidad_minima_de_venta",
    "umv_tipo",
    "categoria_1",
    "categoria_2",
    "categoria_3",
    "categoria_4",
    "aliases",
    "rotacion_index",
    "mental_priority",
    "descripcion",
    "image_url",
    "en_catalogo",
    "is_mock",
    "fuente_hoja",
]

PRICE_FIELDS = [
    "lista_precios_id",
    "nombre",
    "multiplicador_sobre_lista_1",
    "product_code",
    "precio_unidad",
    "is_mock",
]

LISTAS = [
    (1, "Lista Mostrador LGSM", 1.00),
    (2, "Lista Minorista Sugerido", 1.15),
    (3, "Lista Mayorista Especial", 0.90),
    (4, "Lista Gran Distribuidor", 0.85),
]

MARCAS = [
    "MARINARO",
    "SIERRA DE LOS PADRES",
    "ALMA MORA",
    "PUNA",
    "BIANCHI",
    "CAFAYATE",
    "SANTA JULIA",
    "COCA COLA",
    "SPRITE",
    "FANTA",
    "BAGGIO",
    "QUILMES",
    "IMPERIAL",
    "HEINEKEN",
    "STELLA ARTOIS",
    "BRAHMA",
    "TERMA",
    "SEMIX",
    "GORDON",
    "DOBLE V",
    "SERNOVA",
    "IVESS",
    "VERDEFLOR",
    "SIDRA REAL",
    "LUNA NEGRA",
]


def marca_de(nombre: str) -> str:
    up = nombre.upper()
    for m in MARCAS:
        if m in up:
            return m
    if up.startswith("COCA"):
        return "COCA COLA"
    return ""


def familia(nombre: str) -> str:
    u = nombre.upper()
    if "MARINARO" in u:
        return "marinaro"
    if "SIERRA DE LOS PADRES" in u:
        return "sierra_padres"
    if "PUNA" in u and any(x in u for x in ("MALBEC", "TORRONTES", "CABERNET", "SYRAH", "ROSADO", "BLEND", "TEMPRANILLO", "SUAVIGNON", "HELIOS")):
        return "vinos_puna"
    if any(x in u for x in ("QUILMES", "IMPERIAL", "HEINEKEN", "STELLA", "BRAHMA", "ANTARES")):
        return "cervezas"
    if any(x in u for x in ("COCA", "SPRITE", "FANTA")):
        return "gaseosas"
    if "BAGGIO" in u:
        return "jugos"
    if "TERMA" in u:
        return "amargos"
    if "YERBA" in u:
        return "yerba"
    if any(x in u for x in ("PAPA", "SNACK", "CHIZITO", "SEMIX")):
        return "snacks"
    if any(x in u for x in ("SIDRA",)):
        return "sidra"
    if re.search(r"\b(WHISKY|GIN|VODKA|RON|LICOR|FERNET|GINEBRA)\b", u):
        return "blancas"
    if any(x in u for x in ("AGUA MINERAL", "AGUA DE MESA", "AGUA IVESS", "BIDON", "AGUA TONICA")):
        return "aguas"
    if any(
        x in u
        for x in (
            "MALBEC",
            "CABERNET",
            "TORRONTES",
            "TINTO",
            "BLANCO",
            "ROSADO",
            "VINO",
            "PATERA",
            "BORGOÑA",
            "SYRAH",
            "BONARDA",
            "CHARD",
            "MERLOT",
        )
    ):
        return "vinos"
    return "otros"


def categorias(nombre: str) -> tuple[str, str, str, str]:
    u = nombre.upper()
    marca = marca_de(nombre)
    fam = familia(nombre)
    if fam == "marinaro" and any(x in u for x in ("AGUA", "SODA")):
        return ("Bebidas", "Aguas", "Marinaro", "Agua" if "AGUA" in u else "Soda")
    if fam == "marinaro":
        sabor = ""
        for s in ("GRANADINA", "LIMONADA", "MANZANA", "NARANJA", "POMELO", "UVA", "LIMA LIMON", "COLA"):
            if s in u:
                sabor = s.title()
                break
        linea = "Ice" if "ICE" in u else "Gaseosas"
        return ("Bebidas", linea, "Marinaro", sabor)
    mapa = {
        "gaseosas": ("Bebidas", "Gaseosas", marca, ""),
        "cervezas": ("Bebidas", "Cervezas", marca, ""),
        "jugos": ("Bebidas", "Jugos", "Baggio", ""),
        "amargos": ("Bebidas", "Amargos", "Terma", ""),
        "aguas": ("Bebidas", "Aguas", marca, ""),
        "sierra_padres": ("Bebidas", "Aguas", "Sierra de los Padres", ""),
        "blancas": ("Bebidas", "Blancas", marca, ""),
        "sidra": ("Bebidas", "Espumantes", marca, ""),
        "vinos_puna": ("Bebidas", "Vinos", "Puna", ""),
        "vinos": ("Bebidas", "Vinos", marca, ""),
        "snacks": ("Snacks", "Frutos secos" if "MANI" in u or "MANÍ" in u else "Papas", marca, ""),
        "yerba": ("Almacén", "Yerba", marca, ""),
    }
    return mapa.get(fam, ("Bebidas", "Otras", marca, ""))


def parse_formato(nombre: str, fam: str = "") -> tuple[str, str]:
    """Devuelve (n_unidades_texto, contenido) p.ej. ('12', '500 ml')."""
    u = nombre.upper()
    m = re.search(r"(\d+)\s*X\s*([\d,\.]+)(?:\s*(GR|G|ML|LT)\b)?", u)
    if not m:
        return "", ""
    n, vol, unit = m.group(1), m.group(2).replace(".", ","), (m.group(3) or "").upper()
    vol_num = float(vol.replace(",", "."))
    if unit in {"GR", "G"} or fam in {"snacks", "yerba"}:
        return n, f"{vol} g"
    if unit == "ML":
        return n, f"{vol} ml"
    if unit in {"L", "LT"}:
        return n, f"{vol} L"
    if vol_num >= 100:
        return n, f"{vol} ml"
    if vol_num >= 1:
        return n, f"{vol} L"
    return n, f"{vol} ml"


def tipo_envase(nombre: str, fam: str = "") -> str:
    u = nombre.upper()
    if "LATA" in u or "LATON" in u:
        return "latas"
    if "BIDON" in u or "SIFON" in u or "SIFÓN" in u:
        return "sifones" if "SIFON" in u or "SIFÓN" in u else "bidón"
    if "VIDRIO" in u or "BOTELLITA" in u:
        return "botellas de vidrio"
    if fam in {"snacks", "yerba"} or any(x in u for x in ("PAPA", "YERBA", "MANI", "MANÍ")):
        return "unidades"
    return "botellas"


def sabor_de(nombre: str) -> str:
    u = nombre.upper()
    for s, label in (
        ("GRANADINA", "granadina"),
        ("LIMONADA", "limonada"),
        ("LIMA LIMON", "lima limón"),
        ("MANZANA", "manzana"),
        ("NARANJA", "naranja"),
        ("POMELO", "pomelo"),
        ("UVA", "uva"),
        ("COLA", "cola"),
        ("CITRUS", "citrus"),
        ("MALBEC", "malbec"),
        ("TORRONTES", "torrontés"),
        ("CABERNET", "cabernet"),
        ("SYRAH", "syrah"),
        ("TINTO", "tinto"),
        ("BLANCO", "blanco"),
        ("ROSADO", "rosado"),
    ):
        if s in u:
            return label
    return ""


_TIPO = {
    "Gaseosas": "Gaseosa",
    "Aguas": "Agua",
    "Cervezas": "Cerveza",
    "Jugos": "Jugo",
    "Amargos": "Amargo",
    "Blancas": "Bebida blanca",
    "Espumantes": "Sidra",
    "Vinos": "Vino",
    "Ice": "Gaseosa ice",
    "Papas": "Snack",
    "Frutos secos": "Snack",
    "Yerba": "Yerba mate",
    "Otras": "Bebida",
}


def _marca_display(marca: str) -> str:
    if marca == "SIERRA DE LOS PADRES":
        return "Sierra de los Padres"
    if marca == "COCA COLA":
        return "Coca-Cola"
    return marca.title() if marca else ""


def descripcion_corta(nombre: str, cat2: str, qty: int, fam: str = "") -> str:
    marca = marca_de(nombre)
    sabor = sabor_de(nombre)
    n_fmt, vol = parse_formato(nombre, fam)
    envase = tipo_envase(nombre, fam)
    tipo = _TIPO.get(cat2, cat2)
    u = nombre.upper()
    if cat2 == "Aguas" and "SODA" in u:
        tipo = "Soda"
    if cat2 == "Aguas" and "SODA" not in u:
        tipo = "Agua mineral"
    if cat2 == "Cervezas" and "STOUT" in u:
        tipo = "Cerveza stout"
    elif cat2 == "Cervezas":
        tipo = "Cerveza rubia"
    extra = f", marca {_marca_display(marca)}" if marca else ""
    sabor_txt = f" {sabor}" if sabor else ""
    pack = f", caja x{qty}" if qty > 1 else ""
    fmt = f" {envase} de {vol}" if vol else ""
    text = f"{tipo}{sabor_txt}{extra}{pack}{fmt}."
    if len(text.split()) < 10:
        text = text.rstrip(".") + " presentación en pack."
    words = text.split()
    if len(words) > 25:
        text = " ".join(words[:24]) + "."
    return text[0].upper() + text[1:]


def aliases(nombre: str, cat1: str, cat2: str, code: str, qty: int) -> str:
    parts = [nombre.lower(), cat2.lower(), cat1.lower(), code, "caja", "bulto"]
    if qty > 1:
        parts.append(f"caja x{qty}")
        parts.append(f"bulto x{qty}")
    marca = marca_de(nombre)
    if marca:
        parts.append(marca.lower())
        parts.append(f"caja {marca.lower()}")
        parts.append(f"bulto {marca.lower()}")
    sabor = sabor_de(nombre)
    if sabor:
        parts.append(sabor)
        if marca:
            parts.append(f"{marca.lower()} {sabor}")
            parts.append(f"caja {marca.lower()} {sabor}")
    for tok in re.findall(r"[A-ZÁÉÍÓÚÑ0-9]{3,}", nombre):
        if tok.lower() not in {p.lower() for p in parts}:
            parts.append(tok.lower())
    seen: set[str] = set()
    out = []
    for p in parts:
        p = p.strip()
        if not p or p in seen:
            continue
        seen.add(p)
        out.append(p)
    return "|".join(out[:12])


def to_row(art: str, nombre: str, precio: float, qty: int, i: int, n: int) -> dict:
    cat1, cat2, cat3, cat4 = categorias(nombre)
    rot = round(0.93 - (i / max(n - 1, 1)) * 0.50, 2)
    if "MARINARO" in nombre.upper():
        rot = min(0.95, max(rot, 0.82))
    stock = int(80 + rot * 260)
    return {
        "product_code": art,
        "nombre": nombre,
        "precio_lista_1": f"{precio:.2f}",
        "stock": str(stock),
        "unidades_por_bulto": str(qty if qty > 0 else 1),
        "unidad_minima_de_venta": "caja",
        "umv_tipo": "bulto",
        "categoria_1": cat1,
        "categoria_2": cat2,
        "categoria_3": cat3,
        "categoria_4": cat4,
        "aliases": aliases(nombre, cat1, cat2, art, qty),
        "rotacion_index": str(rot),
        "mental_priority": str(round(rot * 0.88, 2)),
        "descripcion": descripcion_corta(nombre, cat2, qty, familia(nombre)),
        "image_url": "",
        "en_catalogo": "true",
        "is_mock": "true",
        "fuente_hoja": "Precios / Mostrador -LGSM",
    }


def write_csv(path: Path, fields: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in fields})


def load_excel() -> list[dict]:
    wb = load_workbook(INP, data_only=True, read_only=True)
    ws = wb["Precios"]
    items: list[dict] = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        art = str(r[4]).strip() if r[4] is not None else ""
        nombre = str(r[5] or "").strip()
        if not art or not nombre:
            continue
        try:
            precio = float(r[15] or 0)
        except (TypeError, ValueError):
            precio = 0
        try:
            qty = int(float(r[9] or 1))
        except (TypeError, ValueError):
            qty = 1
        items.append({"art": art, "nombre": nombre, "precio": precio, "qty": qty, "fam": familia(nombre)})
    wb.close()
    return items


def demo_rank(item: dict) -> tuple:
    """Prioriza formatos que se piden por WhatsApp y sabores distintos."""
    u = item["nombre"].upper().replace(" ", "")
    score = 0
    if "12X500" in u:
        score += 30
    if "6X2,25" in u or "6X2.25" in u:
        score += 24
    if "6X1,5" in u or "6X1.5" in u:
        score += 20
    if "24X473" in u:
        score += 18
    if "QUILMES" in item["nombre"].upper():
        score += 12
    if "CAFAYATE" in item["nombre"].upper():
        score += 10
    if item["fam"] == "blancas":
        if "CAFE" in item["nombre"].upper() or "CAFÉ" in item["nombre"].upper():
            score -= 50
        if any(x in item["nombre"].upper() for x in ("WHISKY", "SMIRNOFF", "GORDON", "CHIVAS")):
            score += 20
    if "6X750" in u:
        score += 16
    if "ICE" in u:
        score += 8
    return (-score, item["nombre"])


def seleccionar(priced: list[dict]) -> list[dict]:
    buckets: dict[str, list[dict]] = defaultdict(list)
    for r in priced:
        buckets[r["fam"]].append(r)
    for fam in buckets:
        buckets[fam].sort(key=demo_rank)

    picked: list[dict] = []
    used: set[str] = set()

    def take(fam: str, n: int) -> None:
        for item in buckets.get(fam, []):
            if len([p for p in picked if p["fam"] == fam]) >= n:
                break
            if item["art"] in used:
                continue
            picked.append(item)
            used.add(item["art"])

    for fam, n in QUOTAS.items():
        take(fam, n)

    if len(picked) < TARGET_MIN:
        leftovers = [r for r in priced if r["art"] not in used]
        leftovers.sort(key=demo_rank)
        for item in leftovers:
            if len(picked) >= TARGET_MIN:
                break
            picked.append(item)
            used.add(item["art"])

    if len(picked) > TARGET_MAX:
        overflow = [p for p in picked if p["fam"] not in {"marinaro", "sierra_padres"}]
        overflow.sort(key=lambda x: (QUOTAS.get(x["fam"], 0), x["nombre"]))
        while len(picked) > TARGET_MAX and overflow:
            drop = overflow.pop(0)
            picked.remove(drop)

    return picked


def main() -> int:
    raw = load_excel()
    priced = [r for r in raw if r["precio"] > 0]
    print(f"[*] origen: {len(raw)} arts, {len(priced)} con Precio Final > 0")

    full_rows: list[dict] = []
    by_code: dict[str, dict] = {}
    for i, r in enumerate(priced):
        row = to_row(r["art"], r["nombre"], r["precio"], r["qty"], i, len(priced))
        full_rows.append(row)
        by_code[row["product_code"]] = row
    write_csv(FULL, PRODUCT_FIELDS, full_rows)
    print(f"[*] catalogo-completo.csv: {len(full_rows)}")

    demo_src = seleccionar(priced)
    n = len(demo_src)
    if not (TARGET_MIN <= n <= TARGET_MAX):
        print(f"[FAIL] recorte fuera de 80–100: {n}")
        return 1

    demo: list[dict] = []
    counts: dict[str, int] = defaultdict(int)
    for i, src in enumerate(demo_src):
        row = dict(by_code[src["art"]])
        row["rotacion_index"] = str(round(0.95 - i * (0.35 / max(n - 1, 1)), 2))
        if src["fam"] == "marinaro":
            row["rotacion_index"] = str(min(0.95, float(row["rotacion_index"]) + 0.08))
        row["mental_priority"] = str(round(float(row["rotacion_index"]) * 0.88, 2))
        row["stock"] = str(int(140 + float(row["rotacion_index"]) * 220))
        row["descripcion"] = descripcion_corta(
            row["nombre"], row["categoria_2"], int(row["unidades_por_bulto"]), src["fam"]
        )
        demo.append(row)
        counts[src["fam"]] += 1

    write_csv(OUT / "phase-01-productos.csv", PRODUCT_FIELDS, demo)
    print(f"[*] phase-01-productos.csv: {len(demo)} SKUs demo")
    print("[*] cupos:", dict(sorted(counts.items(), key=lambda x: -x[1])))

    combined_prices: list[dict] = []
    for list_id, nombre, mult in LISTAS:
        price_rows = []
        for p in demo:
            base = float(p["precio_lista_1"])
            rec = {
                "lista_precios_id": list_id,
                "nombre": nombre,
                "multiplicador_sobre_lista_1": f"{mult:.2f}",
                "product_code": p["product_code"],
                "precio_unidad": f"{round(base * mult, 2):.2f}",
                "is_mock": "true",
            }
            price_rows.append(rec)
            combined_prices.append(rec)
        write_csv(OUT / f"phase-01-lista-precios-{list_id}.csv", PRICE_FIELDS, price_rows)
        print(f"[*] lista {list_id} {nombre}: {len(price_rows)}")
    write_csv(OUT / "phase-01-listas-precios.csv", PRICE_FIELDS, combined_prices)

    print("\n--- RECORTE DEMO ---")
    for p in demo:
        print(
            f"  {p['product_code']:>6}  ${float(p['precio_lista_1']):>10.2f}  "
            f"x{p['unidades_por_bulto']:<3} {p['nombre'][:56]}"
        )
        print(f"          {p['descripcion']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
