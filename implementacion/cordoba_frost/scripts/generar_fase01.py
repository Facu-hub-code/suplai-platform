#!/usr/bin/env python3
"""Fase 1 incremental cordoba_frost: Excel PRODUCTOS-GRAL → CSVs.

schema_name = cordoba_frost (tenant productivo, no demo).
Precio columna Final = lista 1 (Lista General).
No recorta catálogo. is_mock=false.
Los 3 combos de panadería ya existentes se listan pero no se reinsertan.
"""
from __future__ import annotations

import csv
import re
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

SCHEMA = "cordoba_frost"
ROOT = Path(__file__).resolve().parents[1]
INP = ROOT / "inputs" / "PRODUCTOS-GRAL.xlsx"
OUT = ROOT / "outputs"
FULL = ROOT / "inputs" / "catalogo-completo.csv"

EXISTING_SKIP = {
    "COM-COR-01826",
    "COM-COR-01827",
    "COM-COR-01828",
}

# Precio 0 en Excel — no se carga.
SKIP_ZERO_PRICE = {"SIN-CRE-01970"}

PRODUCT_FIELDS = [
    "product_code",
    "nombre",
    "precio_lista_1",
    "stock",
    "unidades_por_bulto",
    "unidad_minima_de_venta",
    "umv_tipo",
    "categoria_1",
    "categoria_2",
    "categoria_3",
    "categoria_4",
    "aliases",
    "rotacion_index",
    "mental_priority",
    "descripcion",
    "image_url",
    "en_catalogo",
    "is_mock",
    "fuente_hoja",
    "es_pesable",
    "sku_origen",
    "nota_sku",
]

PRICE_FIELDS = [
    "lista_precios_id",
    "nombre",
    "multiplicador_sobre_lista_1",
    "product_code",
    "precio_unidad",
    "is_mock",
]

MARCA_PREFIX = {
    "CORDOBA PAN": "COR",
    "CRESFOOD": "CRE",
    "LOMORO": "LOM",
    "PRIPAN-DEVISUR": "PRI",
    "RANCHO ALTO": "RAN",
    "RICCISIMA": "RIC",
    "SAN JOSE": "SJO",
    "INSUMOS HELADERIA": "INS",
    "Q' PIZZA": "QPZ",
    "LA CHURRERIA": "LCH",
    "SOL DE GALICIA": "SOL",
    "GATELIN": "GAT",
    "DRL CONGELADOS": "DRL",
    "QUO": "QUO",
    "FROST CARGO": "FRO",
    "MINYO": "MIN",
    "REBOZADOS": "REB",
    "TEX": "TEX",
}

CAT_PREFIX = {
    "SIN TACC": "SIN",
    "IMPULSIVOS HELADOS IRRESISTIBLES": "IMP",
    "IMPULSIVOS BALDE HELADO 3LTS": "BAL",
    "COMBOS": "COM",
    "PANIFICADOS MEDIALUNAS Y FACTURAS": "PAN",
    "PAPAS CONGELADAS": "PAP",
    "MEDALLONES": "MED",
    "REBOZADOS": "REB",
    "PANIFICADOS PANADERIA": "PAN",
    "PANIFICADOS PASTELERIA": "PAN",
    "INSUMOS COMESTIBLES": "INS",
}


def strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
    )


def clean_name(nombre: str) -> str:
    return re.sub(r"\s+", " ", (nombre or "").strip())


def sanitize_code(raw: str, marca: str) -> tuple[str, str]:
    """Devuelve (codigo_limpio, nota)."""
    code = (raw or "").strip().upper()
    nota = ""
    if not code:
        return "", "sin_codigo_excel"
    original = code
    code = code.replace("'", "").replace("´", "")
    code = re.sub(r"\s+", "", code)
    if code.startswith("-"):
        rest = code[1:]
        marca_u = (marca or "").upper()
        if rest.startswith("PAN-") and "SOL" in marca_u:
            code = "PAN-SOL-" + rest.split("-", 1)[-1]
            nota = f"sanitizado_desde:{original}"
        elif rest.startswith("PAN-"):
            code = "PAN-PAN-" + rest.split("-", 1)[-1]
            nota = f"sanitizado_desde:{original}"
        elif rest.startswith("INS-"):
            code = "INS-INS-" + rest.split("-", 1)[-1]
            nota = f"sanitizado_desde:{original}"
        else:
            code = rest
            nota = f"sanitizado_desde:{original}"
    if "Q" in original and "'" in original:
        # PIZ-Q' -01941 → PIZ-QP-01941
        m = re.match(r"PIZ-Q['\s]*-?(\d+)", original.replace(" ", ""))
        if m:
            code = f"PIZ-QP-{m.group(1)}"
            nota = f"sanitizado_desde:{original}"
    if "LA" in original and " " in original and original.startswith("PAN-LA"):
        m = re.search(r"(\d+)$", original.replace(" ", ""))
        if m:
            code = f"PAN-LCH-{m.group(1)}"
            nota = f"sanitizado_desde:{original}"
    code = re.sub(r"[^A-Z0-9-]", "", code)
    code = re.sub(r"-{2,}", "-", code).strip("-")
    return code, nota


def prefix_for(marca: str, cat: str) -> str:
    cat_u = (cat or "").upper()
    marca_u = (marca or "").upper()
    cat_p = CAT_PREFIX.get(cat_u, "PRD")
    mar_p = MARCA_PREFIX.get(marca_u, "XXX")
    if cat_u.startswith("PANIFICADOS"):
        cat_p = "PAN"
    if cat_u.startswith("IMPULSIVOS"):
        cat_p = "IMP"
    if cat_u.startswith("INSUMOS"):
        cat_p = "INS"
    if cat_u.startswith("BALDES"):
        cat_p = "BAL"
    return f"{cat_p}-{mar_p}"


def next_generated_code(prefix: str, used: set[str], seq: list[int]) -> str:
    while True:
        seq[0] += 1
        code = f"{prefix}-{seq[0]:05d}"
        if code not in used:
            used.add(code)
            return code


def unidades_por_bulto(nombre: str) -> int:
    n = nombre.lower()
    n = re.sub(r"\bunidades\b", "u", n)
    n = re.sub(r"\bunidad(es)?\b", "u", n)
    n = re.sub(r"\bunid\.?\b", "u", n)
    n = n.replace("u.", "u")

    m = re.search(r"x\s*(\d+)\s*paquetes[^\d]*(\d+)\s*u", n)
    if m:
        return int(m.group(1)) * int(m.group(2))

    m = re.search(r"(\d+)\s*doc", n)
    if m:
        return int(m.group(1)) * 12

    # Quitar pesos/volúmenes para no tomar 120 de "120 gr" ni 10 de "10.5 kg".
    n_wo_w = re.sub(r"\d+(?:[.,]\d+)?\s*(?:grs?|g|kg|lts?|l|cc|ml)\b", " ", n)

    patterns = [
        r"x\s*(\d+)\s*u\b",
        r"(\d+)\s*u\b",
        r"pack\s*x?\s*(\d+)",
        r"x\s*(\d+)\s*paquetes",
        r"caja\s*x\s*(\d+)\b",
    ]
    for pat in patterns:
        m = re.search(pat, n_wo_w)
        if m:
            val = int(m.group(1))
            if 2 <= val <= 500:
                return val
    return 1


def es_pesable(nombre: str) -> bool:
    n = nombre.lower()
    if re.search(r"\bx\s*kg\b", n) and not re.search(r"\d+\s*u\b", n):
        return True
    return False


def cat_levels(marca: str, cat_excel: str, nombre: str) -> tuple[str, str, str, str]:
    cat = (cat_excel or "").strip() or "General"
    n = nombre.lower()
    cat_u = cat.upper()

    if cat_u.startswith("PANIFICADOS") or "MEDIALUNA" in cat_u or "FACTURA" in cat_u:
        c1 = "Panificados"
    elif cat_u == "COMBOS":
        c1 = "Combos"
    elif any(
        k in cat_u
        for k in (
            "IMPULSIVOS",
            "BALDES",
            "TORTAS HELADAS",
            "HELADO",
        )
    ):
        c1 = "Helados"
    elif cat_u.startswith("INSUMOS"):
        c1 = "Insumos"
    elif cat_u == "SIN TACC":
        c1 = "Sin TACC"
    elif cat_u in {
        "REBOZADOS",
        "MEDALLONES",
        "HAMBURGUESAS",
        "PAPAS CONGELADAS",
        "FRUTAS CONGELADAS",
        "PIZZAS Y EMPANADAS",
    }:
        c1 = "Congelados"
    elif cat_u == "PRODUCTOS AFA":
        c1 = "Helados"
    else:
        c1 = "General"

    c2 = cat
    if not cat_excel:
        if "mani" in n or "rocklets" in n or "obleas" in n:
            c2 = "Insumos Comestibles"
            c1 = "Insumos"
        else:
            c2 = "Insumos Comestibles"
            c1 = "Insumos"

    c3 = (marca or "").strip() or "Sin marca"

    c4 = "Caja"
    if "combo" in n:
        c4 = "Combo"
    elif re.search(r"\bvaso\b", n):
        c4 = "Vaso"
    elif "palito" in n:
        c4 = "Palito"
    elif "balde" in n:
        c4 = "Balde"
    elif "pote" in n:
        c4 = "Pote"
    elif "pack" in n:
        c4 = "Pack"
    elif re.search(r"\bkg\b", n):
        c4 = "Kg"
    elif "unidad" in n or re.search(r"x\s*1\b", n):
        c4 = "Unidad"
    return c1, c2, c3, c4


def rotacion(marca: str, cat1: str, cat2: str) -> float:
    m = (marca or "").upper()
    c2 = (cat2 or "").upper()
    if "COMBO" in c2:
        return 0.88
    if m in {"LOMORO", "CORDOBA PAN", "GATELIN"}:
        return 0.82
    if m in {"CRESFOOD", "DRL CONGELADOS", "QUO", "SOL DE GALICIA", "PRIPAN-DEVISUR"}:
        return 0.62
    if cat1 == "Insumos":
        return 0.35
    if m in {"RICCISIMA", "RANCHO ALTO", "SAN JOSE", "TEX", "Q' PIZZA"}:
        return 0.55
    return 0.40


def stock_for(rot: float) -> int:
    if rot >= 0.75:
        return 250
    if rot >= 0.55:
        return 120
    if rot >= 0.4:
        return 60
    return 30


def tipo_sustantivo(cat2: str, nombre: str) -> str:
    n = nombre.lower()
    c = (cat2 or "").lower()
    if "combo" in n or c == "combos":
        return "Combo"
    if "baguett" in n:
        return "Baguettín congelado"
    if "chipa" in n:
        return "Chipá congelada"
    if "criollo" in n:
        return "Criollo congelado"
    if "factura" in n or "persianita" in n:
        return "Factura congelada"
    if "medialuna" in n:
        return "Medialuna congelada"
    if "pan " in n or n.startswith("pan"):
        return "Pan congelado"
    if "pizza" in n:
        return "Pizza congelada"
    if "empanada" in n:
        return "Empanada congelada"
    if "palito" in n and "agua" in c:
        return "Palito de agua"
    if "palito" in n:
        return "Palito de crema"
    if "vaso" in n:
        return "Helado en vaso"
    if "torta" in n:
        return "Torta helada"
    if "postre" in n or "alfajor" in n or "bombon" in n or "bombón" in n:
        return "Postre helado"
    if "balde" in n:
        return "Balde de helado"
    if "salsa" in n:
        return "Salsa para heladería"
    if "cono" in n or "cucurucho" in n:
        return "Cono para helado"
    if "vaso pasta" in n or "envase" in n or "telgopor" in n:
        return "Envase descartable"
    if "papa" in n:
        return "Papa congelada"
    if "medallon" in n or "medallón" in n:
        return "Medallón congelado"
    if "milanesa" in n or "reboz" in n:
        return "Rebozado congelado"
    if "hamburg" in n or "bife" in n:
        return "Hamburguesa congelada"
    if "fruta" in c or "fruta" in n:
        return "Fruta congelada"
    if "dona" in n or "croissant" in n or "churro" in n:
        return "Pastelería congelada"
    if "sin tacc" in c or "sin tacc" in n:
        return "Producto sin TACC"
    if cat2.startswith("Insumos"):
        return "Insumo de heladería"
    if cat2.startswith("Baldes") or "10lt" in c:
        return "Balde de helado 10 L"
    return "Producto congelado"


def descripcion(nombre: str, marca: str, cat2: str, upb: int) -> str:
    tipo = tipo_sustantivo(cat2, nombre)
    marca_txt = (marca or "").strip() or "sin marca"
    n = clean_name(nombre)
    extra = []
    m_peso = re.search(
        r"(\d+(?:[.,]\d+)?\s*(?:grs?|g|kg|lts?|l|cc|ml))\b", n, re.I
    )
    if m_peso:
        extra.append(m_peso.group(1).replace("grs", "g").replace("GRS", "g"))
    if upb > 1:
        extra.append(f"x{upb} unidades")
    elif re.search(r"\bkg\b", n.lower()):
        extra.append("venta por kg")
    bits = [tipo, n.rstrip("."), f"marca {marca_txt}"]
    if extra:
        bits.append(", ".join(extra))
    text = ", ".join(bits) + "."
    words = text.split()
    if len(words) > 25:
        text = " ".join(words[:25]).rstrip(",.") + "."
    if len(text.split()) < 10:
        text = (
            f"{tipo} {n.rstrip('.')}, marca {marca_txt}, "
            f"categoría {cat2 or 'general'}, presentación de catálogo."
        )
        words = text.split()
        if len(words) > 25:
            text = " ".join(words[:25]).rstrip(",.") + "."
    return text


def aliases(nombre: str, marca: str, cat2: str, upb: int) -> str:
    n = clean_name(nombre)
    parts = [
        n,
        n.lower(),
        strip_accents(n).lower(),
        (marca or "").strip(),
        cat2 or "",
    ]
    short = re.sub(r"\b(congelad[ao]s?|caja|pack|x\s*\d+u\.?)\b", "", n, flags=re.I)
    short = re.sub(r"\s+", " ", short).strip(" -")
    if short and short.lower() != n.lower():
        parts.append(short)
    if upb > 1:
        parts.extend(["caja", "bulto", f"caja x{upb}"])
    if "combo" in n.lower():
        parts.append("combo")
    if "helado" in (cat2 or "").lower() or "impulsivo" in (cat2 or "").lower():
        parts.extend(["helado", "helados"])
    if "palito" in n.lower():
        parts.append("palito")
    if "balde" in n.lower():
        parts.append("balde")
    if "medialuna" in n.lower():
        parts.extend(["medialuna", "medialunas"])
    seen: set[str] = set()
    out: list[str] = []
    for p in parts:
        p = re.sub(r"\s+", " ", (p or "").strip())
        if not p:
            continue
        key = strip_accents(p).lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(p)
    return "|".join(out[:12])


def parse_precio(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def load_rows() -> list[dict]:
    wb = load_workbook(INP, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row) + [None] * 7
        _, marca, cat, nombre, codigo, precio, _ = vals[:7]
        nombre = clean_name(str(nombre) if nombre else "")
        if not nombre:
            continue
        rows.append(
            {
                "excel_row": i,
                "marca": clean_name(str(marca) if marca else ""),
                "cat": clean_name(str(cat) if cat else ""),
                "nombre": nombre,
                "codigo_raw": str(codigo).strip() if codigo else "",
                "precio": parse_precio(precio),
            }
        )
    wb.close()
    return rows


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    raw_rows = load_rows()

    used: set[str] = set()
    # Reservar códigos existentes en BD que no están en el Excel.
    used.update({"COM-HEL-INICIAL", "COM-HEL-MEDIO", "COM-HEL-PREMIUM", "ENVIO-DOM"})
    used.update(EXISTING_SKIP)

    seq = [20000]
    prepared: list[dict] = []
    skipped: list[str] = []

    # Primera pasada: sanitizar códigos presentes.
    sanitized: list[dict] = []
    counts: Counter[str] = Counter()
    for r in raw_rows:
        code, nota = sanitize_code(r["codigo_raw"], r["marca"])
        sanitized.append({**r, "code": code, "nota": nota})
        if code:
            counts[code] += 1

    # Segunda: asignar códigos faltantes / duplicados / colisiones.
    seen_once: set[str] = set()
    for r in sanitized:
        precio = r["precio"]
        if precio is None or precio <= 0:
            skipped.append(
                f"fila {r['excel_row']} {r['nombre']}: precio={precio} (omitido)"
            )
            continue
        code = r["code"]
        nota = r["nota"]
        origen = r["codigo_raw"] or ""

        if not code:
            prefix = prefix_for(r["marca"], r["cat"])
            code = next_generated_code(prefix, used, seq)
            nota = "generado_sin_codigo_excel"
        elif code in SKIP_ZERO_PRICE:
            skipped.append(f"{code} precio 0 omitido")
            continue
        elif counts[code] > 1:
            if code in seen_once:
                prefix = prefix_for(r["marca"], r["cat"])
                new_code = next_generated_code(prefix, used, seq)
                nota = f"duplicado_excel:{code}"
                origen = code
                code = new_code
            else:
                seen_once.add(code)
        if code in used and nota.startswith("sanitizado"):
            prefix = prefix_for(r["marca"], r["cat"])
            new_code = next_generated_code(prefix, used, seq)
            nota = f"{nota}|colision:{code}"
            code = new_code
        used.add(code)

        upb = unidades_por_bulto(r["nombre"])
        c1, c2, c3, c4 = cat_levels(r["marca"], r["cat"], r["nombre"])
        rot = rotacion(r["marca"], c1, c2)
        prepared.append(
            {
                "product_code": code,
                "nombre": r["nombre"],
                "precio_lista_1": int(precio) if precio == int(precio) else precio,
                "stock": stock_for(rot),
                "unidades_por_bulto": upb,
                "unidad_minima_de_venta": "unidad",
                "umv_tipo": "unidad",
                "categoria_1": c1,
                "categoria_2": c2,
                "categoria_3": c3,
                "categoria_4": c4,
                "aliases": aliases(r["nombre"], r["marca"], c2, upb),
                "rotacion_index": f"{rot:.2f}",
                "mental_priority": "0.0",
                "descripcion": descripcion(r["nombre"], r["marca"], c2, upb),
                "image_url": "",
                "en_catalogo": "true",
                "is_mock": "false",
                "fuente_hoja": "Productos | General",
                "es_pesable": "true" if es_pesable(r["nombre"]) else "false",
                "sku_origen": origen,
                "nota_sku": nota,
                "excel_row": r["excel_row"],
            }
        )

    with FULL.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=PRODUCT_FIELDS, extrasaction="ignore")
        w.writeheader()
        w.writerows(prepared)

    to_insert = [p for p in prepared if p["product_code"] not in EXISTING_SKIP]
    already = [p for p in prepared if p["product_code"] in EXISTING_SKIP]
    generated = [p for p in to_insert if p["nota_sku"].startswith("generado") or p["nota_sku"].startswith("duplicado")]
    sanitized_rows = [p for p in to_insert if p["nota_sku"].startswith("sanitizado")]

    prod_path = OUT / "phase-01-productos-gral.csv"
    with prod_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=PRODUCT_FIELDS, extrasaction="ignore")
        w.writeheader()
        w.writerows(to_insert)

    price_path = OUT / "phase-01-lista-precios-1-gral.csv"
    with price_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=PRICE_FIELDS)
        w.writeheader()
        for p in to_insert:
            w.writerow(
                {
                    "lista_precios_id": 1,
                    "nombre": "Lista General",
                    "multiplicador_sobre_lista_1": 1.00,
                    "product_code": p["product_code"],
                    "precio_unidad": p["precio_lista_1"],
                    "is_mock": "false",
                }
            )

    notes = ROOT / "outputs" / "phase-01-gral-notas.md"
    lines = [
        "# Córdoba Frost — alta catálogo PRODUCTOS GRAL",
        "",
        f"- Schema: `{SCHEMA}`",
        f"- Filas Excel con nombre: {len(raw_rows)}",
        f"- Con precio > 0: {len(prepared)}",
        f"- Ya existentes (no se reinsertan): {len(already)} — {', '.join(p['product_code'] for p in already)}",
        f"- A insertar: {len(to_insert)}",
        f"- SKUs generados o desduplicados: {len(generated)}",
        f"- SKUs sanitizados: {len(sanitized_rows)}",
        "",
        "## Omitidos",
        "",
    ]
    for s in skipped:
        lines.append(f"- {s}")
    lines += ["", "## SKUs generados / duplicados", ""]
    for p in generated:
        lines.append(
            f"- `{p['product_code']}` ← origen `{p['sku_origen'] or '(vacío)'}` "
            f"({p['nota_sku']}) — {p['nombre']}"
        )
    lines += ["", "## SKUs sanitizados", ""]
    for p in sanitized_rows:
        lines.append(
            f"- `{p['product_code']}` ← `{p['sku_origen']}` — {p['nombre']}"
        )
    notes.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"[*] universo {len(prepared)} → {FULL}")
    print(f"[*] insertar {len(to_insert)} → {prod_path}")
    print(f"[*] precios lista 1 → {price_path}")
    print(f"[*] notas → {notes}")
    print(f"[*] ya existentes: {len(already)}")
    print(f"[*] generados/dup: {len(generated)} sanitizados: {len(sanitized_rows)}")
    print(f"[*] omitidos: {len(skipped)}")
    by_c1 = Counter(p["categoria_1"] for p in to_insert)
    print("[*] categoria_1:")
    for k, v in by_c1.most_common():
        print(f"    {v:4d}  {k}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
