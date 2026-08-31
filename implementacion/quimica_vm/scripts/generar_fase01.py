#!/usr/bin/env python3
"""Genera CSVs Fase 1 para quimica_vm desde liquidos.xlsx (lista revendedor sep 2026).

Curado: ~80 SKUs representativos (una presentación por código, preferir 5 L).
Precios SIN IVA. schema_name = quimica_vm.
"""
from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "inputs" / "liquidos.xlsx"
OUT = ROOT / "outputs"

# 80 códigos: cubren las 10 familias de la lista, sin explotar bidón 200/1000 L.
INCLUDE = {
    1001, 1003, 1005, 1007, 1008, 1009, 1010, 1012, 1013, 1014, 1016, 1017, 1019, 1021,
    2001, 2005, 2008, 2011, 2014, 2016, 2019, 2022,
    3006, 3011, 3014, 3017, 3019, 3020, 3025, 3026,
    4001, 4002, 4003, 4005, 4006, 4007,
    5001, 5002, 5003, 5004, 5006, 5009, 5010, 5011,
    6002, 6003, 6005, 6006, 6007, 6008, 6009, 6010, 6011, 6013, 6015,
    7001, 7007, 7008, 7010, 7012, 7014, 7015,
    8001, 8002, 8003, 8004, 8005, 8008, 8010,
    9003, 9005, 9006, 9008, 9018, 9026, 9028, 9029, 9030, 9031,
    10001, 10002,
}

LISTAS = [
    (1, "Lista 1", "Lista revendedor septiembre 2026 (sin IVA)", 1.00),
    (2, "Lista 2", "Lista Minorista Sugerido", 1.15),
    (3, "Lista 3", "Lista Mayorista Especial", 0.90),
    (4, "Lista 4", "Lista Gran Distribuidor", 0.85),
]

PRODUCT_FIELDS = [
    "product_code",
    "nombre",
    "precio_lista_1",
    "stock",
    "unidades_por_bulto",
    "unidad_minima_de_venta",
    "umv_tipo",
    "categoria_1",
    "categoria_2",
    "categoria_3",
    "categoria_4",
    "aliases",
    "rotacion_index",
    "mental_priority",
    "descripcion",
    "image_url",
    "en_catalogo",
    "is_mock",
    "fuente_hoja",
]

CAT1 = {
    1000: "Ropa",
    2000: "Pisos",
    3000: "Desodorantes",
    4000: "Detergentes",
    5000: "Cloro e insecticidas",
    6000: "Desengrasantes y limpiadores",
    7000: "Automotor",
    8000: "Pileta",
    9000: "Varios",
    10000: "Envases",
}

HIGH_ROT = (
    "JABON LIQUIDO",
    "SUAVIZANTE",
    "DETERGENTE",
    "LAVANDINA",
    "DESENGRASANTE",
    "CLORO PURO",
    "DESTAPA",
)

MED_ROT = (
    "CERA",
    "ALGUICIDA",
    "SHAMPOO",
    "PERFUME",
    "LIMPIA",
    "INSECTICIDA",
    "VINAGRE",
)


def norm(s: str) -> str:
    return unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode("ascii").upper()


def family_code(code: int) -> int:
    if code >= 10000:
        return 10000
    return (code // 1000) * 1000


def pick_presentation(name: str, children: list[tuple[str, float]]) -> tuple[str, float] | None:
    """Prefiere 5 L (mínimo mayorista), luego 1 L/1 kg, luego precio en la fila del código."""
    scored: list[tuple[int, str, float]] = []
    for label, price in children:
        lab = norm(label)
        score = 50
        if re.search(r"\bX\s*5\s*(LITROS?|LT|L)\b", lab) or "X 5 LITROS" in lab:
            score = 10
        elif "1LT PARA 80" in lab or "1 LT PARA 80" in lab:
            score = 15
        elif re.search(r"\bX\s*1\s*(LITROS?|LT|L|KG)\b", lab) or "X 1 LITRO" in lab or "X 1KG" in lab or lab.endswith("X 1LT"):
            score = 20
        elif re.search(r"\bX\s*20\s*(LITROS?|LT)\b", lab):
            score = 22
        elif "BOTELLA DE 1LT" in lab and "BIDON" not in lab:
            score = 25
        elif "TAMBOR" in lab or "TANQUE" in lab or "200 LITROS" in lab or "1000 LITROS" in lab:
            score = 90
        elif "25KG" in lab:
            score = 80
        scored.append((score, label, price))
    if not scored:
        return None
    scored.sort(key=lambda x: x[0])
    _, label, price = scored[0]
    return label, price


def pretty_pack(label: str) -> str:
    lab = " ".join(str(label).split())
    if lab.startswith("*") and "PRECIO" not in norm(lab) and "LITRO" not in norm(lab):
        return ""
    mapping = [
        (r"(?i)x\s*5\s*litros?(?:\s*\(.*\))?", "5 L"),
        (r"(?i)x\s*10\s*litros?(?:\s*\(.*\))?", "10 L"),
        (r"(?i)x\s*20\s*litros?(?:\s*\(.*\))?", "20 L"),
        (r"(?i)x\s*1\s*litro", "1 L"),
        (r"(?i)x\s*1\s*lt", "1 L"),
        (r"(?i)1lt para 80lt", "1 L (rinde 80 L)"),
        (r"(?i)x\s*1lt", "1 L"),
        (r"(?i)x\s*1\s*kg", "1 kg"),
        (r"(?i)x\s*500gr", "500 g"),
        (r"(?i)botella de 1lt$", "botella 1 L"),
        (r"(?i)para bid[oó]n", "a granel / bidón"),
    ]
    for pat, out in mapping:
        if re.search(pat, lab):
            extra = ""
            if "INCLUIDO" in norm(lab) or "INCLUIDO" in norm(lab):
                extra = " (bidón incluido)"
            if "NO INCLUYE" in norm(lab):
                extra = " (sin bidón)"
            return out + extra
    return lab


def noun_and_cats(name: str, cat1: str) -> tuple[str, str, str, str]:
    n = norm(name)
    if "BIDON" in n:
        return "Bidón plástico", "Envases", "Bidones", "5–10 L"
    if "BOTELLA" in n and "PET" in n:
        return "Botella PET", "Envases", "Botellas", "PET"
    if "JABON LIQUIDO" in n or "JABON EXTRA" in n:
        return "Jabón líquido para ropa", cat1, "Jabón líquido", "Soap Clean"
    if "SUAVIZANTE" in n:
        return "Suavizante para ropa", cat1, "Suavizantes", "Soap Clean"
    if "ADITIVO PERFUMANTE" in n:
        return "Aditivo perfume para lavandina", cat1, "Lavandina", "Aditivo"
    if "JABON TOCADOR" in n:
        return "Jabón líquido de tocador", cat1, "Higiene personal", "Fragancia"
    if "LAVANDINA ROPA" in n or (n.startswith("LAVANDINA") and "GEL" not in n and "50" not in n):
        return "Lavandina para ropa", cat1, "Lavandina", "Ropa"
    if "LAVANDINA 50" in n or "CLORO PURO" in n:
        return "Lavandina / cloro", cat1, "Cloro", "Desinfectante"
    if "LAVANDINA EN GEL" in n:
        return "Lavandina en gel", cat1, "Baño y cocina", "Gel"
    if "DETERGENTE" in n:
        return "Detergente para vajilla", cat1, "Detergentes", "Cocina"
    if "DESENGRASANTE" in n and "MOTOR" in n:
        return "Desengrasante de motor", cat1, "Taller", "Motor"
    if "DESENGRASANTE" in n and ("CARROCERIA" in n or "CARCLEAN" in n or "LAVA CARROCERIA" in n):
        return "Desengrasante lava carrocería", cat1, "Lavadero", "Carrocería"
    if "DESENGRASANTE" in n:
        return "Desengrasante", cat1, "Cocina e industrial", "Desengrasantes"
    if "REMOVEDOR" in n:
        return "Removedor de ceras", cat1, "Mantenimiento", "Removedor"
    if "CERA" in n:
        return "Cera para pisos", cat1, "Ceras", "Autobrillo"
    if "CURADOR" in n:
        return "Curador de piso", cat1, "Ceras", "Curador"
    if "LAMPAZO" in n:
        return "Líquido de lampazo", cat1, "Mantenimiento", "Lampazo"
    if "BRILLO AL BALDE" in n:
        return "Brillo al balde", cat1, "Mantenimiento", "Pisos"
    if "DESODORANTE" in n:
        return "Desodorante de ambientes concentrado", cat1, "Ambientadores", "1+80"
    if "PERFUME ROPA" in n:
        return "Perfume para ropa", cat1, "Perfume textil", "Concentrado"
    if "PERFUME AUTO" in n or "PERFUME AMBIENTAL" in n:
        return "Perfume ambiental", cat1, "Ambientadores", "1 L"
    if "QUITAMANCHA" in n:
        return "Quitamanchas", cat1, "Tratamiento", "Ropa"
    if "CAMELLITO" in n:
        return "Jabón camellito", cat1, "Jabón líquido", "Pretatamiento"
    if "ALGUICIDA" in n:
        return "Alguicida para pileta", cat1, "Alguicidas", "Pool"
    if "CLARIFICANTE" in n:
        return "Clarificante de pileta", cat1, "Clarificantes", "Pool"
    if "CLORO GRANULADO" in n or "PASTILLA DE CLORO" in n:
        return "Cloro para pileta", cat1, "Cloro", "Granulado/pastilla"
    if "REGULADOR" in n and "PH" in n:
        return "Regulador de pH para pileta", cat1, "pH", "Pool"
    if "INSECTICIDA" in n or "DERRIBANTE" in n or "CREOLINA" in n:
        return "Insecticida / derribante", cat1, "Insecticidas", "Concentrado"
    if "REPELENTE" in n:
        return "Repelente", cat1, "Repelentes", "Hogar"
    if "DESTAPA" in n:
        return "Destapa cañerías", cat1, "Baño", "Cañerías"
    if "LIMPIA INODOROS" in n:
        return "Limpia inodoros", cat1, "Baño", "Inodoro"
    if "LIMPIA HORNOS" in n:
        return "Limpia hornos", cat1, "Cocina", "Horno"
    if "LIMPIA ALFOMBRAS" in n:
        return "Limpia alfombras y tapizados", cat1, "Textil", "Alfombras"
    if "LIMPIAVIDRIO" in n:
        return "Limpiavidrios concentrado", cat1, "Vidrios", "Concentrado"
    if "LUSTRAMUEBLES" in n:
        return "Lustramuebles", cat1, "Muebles", "Cera"
    if "QUITASARRO" in n:
        return "Quitasarro líquido", cat1, "Baño", "Sarro"
    if "SHAMPOO PARA AUTO" in n:
        return "Shampoo para auto", cat1, "Lavadero", "Shampoo"
    if "SILICONA" in n or "RENOVADOR DE CAUCHO" in n:
        return "Renovador automotor", cat1, "Interiores", "Silicona/caucho"
    if "LIMPIA PARABRISAS" in n:
        return "Limpia parabrisas", cat1, "Lavadero", "Vidrios"
    if "SHAMPOO PARA PERROS" in n:
        return "Shampoo para perros", cat1, "Mascotas", " Antipulgas"
    if "SHAMPOO PARA CABELLOS" in n:
        return "Shampoo capilar", cat1, "Higiene personal", "Cabello"
    if "ALCOHOL" in n:
        return "Alcohol", cat1, "Solventes", "Alcohol"
    if "BICARBONATO" in n:
        return "Bicarbonato de sodio", cat1, "Polvos", "Soap Clean"
    if "PERCARBONATO" in n:
        return "Percarbonato de sodio", cat1, "Polvos", "Soap Clean"
    if "ACIDO CITRICO" in n or "CIDO CITRICO" in n:
        return "Ácido cítrico en polvo", cat1, "Polvos", "Soap Clean"
    if "VINAGRE" in n:
        return "Vinagre de limpieza", cat1, "Ácidos suaves", "Limpieza"
    return "Líquido de limpieza", cat1, cat1, "Línea propia"


def marca(name: str) -> str:
    n = norm(name)
    if "POOL CLEAN" in n:
        return "Pool Clean"
    if "CARCLEAN" in n:
        return "CarClean"
    if "SOAP CLEAN" in n:
        return "Soap Clean"
    if "QUIMICA VM" in n or "QVM" in n:
        return "Química VM"
    return "Química VM"


def tipo_equivalente(name: str) -> str:
    m = re.search(r"TIPO\s+([^)\"]+)", name, re.I)
    if m:
        return m.group(1).strip(" .")
    return ""


def dilucion(name: str) -> str:
    n = name
    if re.search(r"L\.?P\.?U", n, re.I):
        return "listo para usar"
    m = re.search(r"1\s*\+\s*(\d+)", n)
    if m:
        return f"concentrado 1+{m.group(1)} (1 L de producto + {m.group(1)} L de agua)"
    m = re.search(r"\((\d+(?:[.,]\d+)?)\s*L?\s*\+\s*(\d+(?:[.,]\d+)?)\s*LT", n, re.I)
    if m:
        return f"concentrado ({m.group(1)} L + {m.group(2)} L de agua)"
    m = re.search(r"1LT EN ([\d.]+)", n, re.I)
    if m:
        return f"dosis 1 L en {m.group(1)} L de agua"
    m = re.search(r"1 EN ([\d.]+)", n, re.I)
    if m:
        return f"dosis 1 L en {m.group(1)} L de agua"
    return ""


def descripcion(noun: str, name: str, pack: str, price_unit: str) -> str:
    brand = marca(name)
    eq = tipo_equivalente(name)
    dil = dilucion(name)
    bits = [noun]
    if eq:
        bits.append(f"tipo {eq}")
    bits.append(f"marca {brand}")
    if pack:
        bits.append(pack)
    if dil:
        bits.append(dil)
    if "PRECIO POR LITRO" in norm(name) or price_unit == "litro":
        bits.append("precio por litro")
    text = ", ".join(bits) + "."
    words = text.split()
    if len(words) > 25:
        text = " ".join(words[:24]) + "."
    if len(words) < 10:
        text = f"{noun}, marca {brand}, {pack or 'venta a granel'}, línea Química VM Villa María."
    return text


def aliases(code: int, name: str, noun: str, pack: str) -> str:
    parts = [str(code), noun.lower(), name.split('"')[0].strip()[:40]]
    brand = marca(name)
    parts.append(brand)
    eq = tipo_equivalente(name)
    if eq:
        parts.append(eq)
    if pack:
        parts.append(pack)
    if re.search(r"L\.?P\.?U", name, re.I):
        parts += ["LPU", "listo para usar"]
    n = norm(name)
    if "LAVANDINA" in n:
        parts.append("cloro")
    if "JABON LIQUIDO" in n:
        parts.append("detergente ropa")
    seen: list[str] = []
    for p in parts:
        p = " ".join(str(p).split())
        if p and p.lower() not in {x.lower() for x in seen}:
            seen.append(p)
    return "|".join(seen[:8])


def rotacion(name: str) -> tuple[float, int]:
    n = norm(name)
    if any(k in n for k in HIGH_ROT):
        return 0.88, 220
    if any(k in n for k in MED_ROT):
        return 0.62, 120
    if "BIDON" in n or "BOTELLA" in n:
        return 0.45, 80
    return 0.35, 60


def parse_workbook() -> list[dict]:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    products: dict[int, dict] = {}
    current_code: int | None = None
    last_family: int | None = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3, values_only=True):
        a, b, c = row
        if isinstance(a, (int, float)) and a == int(a):
            code = int(a)
            name = str(b or "").strip()
            if str(c).strip().upper() == "PRECIO" or (code % 1000 == 0 and code >= 1000):
                current_code = None
                last_family = code
                shared_children = []
                continue
            current_code = code
            products[code] = {
                "code": code,
                "name": name,
                "row_price": c if isinstance(c, (int, float)) else None,
                "children": [],
                "family": last_family or family_code(code),
            }
            continue
        if current_code and b and isinstance(c, (int, float)):
            products[current_code]["children"].append((str(b).strip(), float(c)))
    shared_3000: list[tuple[str, float]] = []
    grab = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3, values_only=True):
        a, b, c = row
        if isinstance(a, (int, float)) and int(a) == 3000:
            grab = True
            continue
        if grab and isinstance(a, (int, float)) and int(a) >= 3001:
            grab = False
        if grab and b and isinstance(c, (int, float)):
            shared_3000.append((str(b).strip(), float(c)))
    for p in products.values():
        if family_code(p["code"]) == 3000 and not p["children"] and p["row_price"] is None:
            p["children"] = list(shared_3000)
    return [p for p in products.values() if p["code"] in INCLUDE]


def main() -> None:
    rows_in = parse_workbook()
    found = {p["code"] for p in rows_in}
    missing = sorted(INCLUDE - found)
    if missing:
        raise SystemExit(f"Faltan códigos en el Excel: {missing}")

    out_rows: list[dict] = []
    for p in sorted(rows_in, key=lambda x: x["code"]):
        children = list(p["children"])
        if p["row_price"] is not None:
            children.append(("", float(p["row_price"])))
        picked = pick_presentation(p["name"], children)
        if not picked:
            raise SystemExit(f"Sin precio: {p['code']} {p['name']}")
        label, price = picked
        pack = pretty_pack(label) if label else ""
        cat1 = CAT1[family_code(p["code"])]
        noun, c1, c2, c3 = noun_and_cats(p["name"], cat1)
        price_unit = "litro"
        if any(k in norm(p["name"] + " " + label) for k in ("BOTELLA", "POTE", "PASTILLA", "BIDON DE", "ENVASADA", "X200GR", "X 1KG", "X 500")):
            if "BIDON DE" in norm(p["name"]) or "BOTELLA PLASTICA" in norm(p["name"]) or "PASTILLA" in norm(p["name"]):
                price_unit = "unidad"
            elif "X 1KG" in norm(label) or "X 500" in norm(label):
                price_unit = "kg"
        nombre = p["name"]
        if pack and pack.lower() not in nombre.lower():
            nombre = f"{nombre} — {pack}"
        rot, stock = rotacion(p["name"])
        desc = descripcion(noun, p["name"], pack, price_unit)
        out_rows.append(
            {
                "product_code": str(p["code"]),
                "nombre": nombre,
                "precio_lista_1": f"{price:.2f}",
                "stock": str(stock),
                "unidades_por_bulto": "1",
                "unidad_minima_de_venta": "unidad",
                "umv_tipo": "unidad",
                "categoria_1": c1,
                "categoria_2": c2,
                "categoria_3": c3,
                "categoria_4": pack or "granel",
                "aliases": aliases(p["code"], p["name"], noun, pack),
                "rotacion_index": f"{rot:.2f}",
                "mental_priority": f"{rot:.2f}",
                "descripcion": desc,
                "image_url": "",
                "en_catalogo": "true",
                "is_mock": "false",
                "fuente_hoja": "SEPTIEMBRE 2026",
            }
        )

    OUT.mkdir(exist_ok=True)
    prod_path = OUT / "phase-01-productos.csv"
    with prod_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=PRODUCT_FIELDS)
        w.writeheader()
        w.writerows(out_rows)

    for lid, lname, ldesc, mult in LISTAS:
        path = OUT / f"phase-01-lista-precios-{lid}.csv"
        with path.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["product_code", "precio_unidad", "is_mock"])
            w.writeheader()
            for r in out_rows:
                base = float(r["precio_lista_1"])
                w.writerow(
                    {
                        "product_code": r["product_code"],
                        "precio_unidad": f"{round(base * mult, 2):.2f}",
                        "is_mock": "false" if lid == 1 else "true",
                    }
                )

    lists_path = OUT / "phase-01-listas-precios.csv"
    with lists_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["id", "nombre", "descripcion", "multiplicador", "is_mock"])
        w.writeheader()
        for lid, lname, ldesc, mult in LISTAS:
            w.writerow(
                {
                    "id": lid,
                    "nombre": lname,
                    "descripcion": ldesc,
                    "multiplicador": f"{mult:.2f}",
                    "is_mock": "false" if lid == 1 else "true",
                }
            )

    print(f"OK {len(out_rows)} productos → {prod_path}")
    by_cat: dict[str, int] = {}
    for r in out_rows:
        by_cat[r["categoria_1"]] = by_cat.get(r["categoria_1"], 0) + 1
    for k, v in sorted(by_cat.items()):
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
