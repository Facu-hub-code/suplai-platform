#!/usr/bin/env python3
"""Genera CSVs de Fase 1 (catálogo) para benavidez desde Especias_Benavidez_Precios.xlsx.

Precio del Excel = Precio (ARS) de la lista B2B pública (precios finales con IVA).
schema_name = benavidez (confirmado).
"""
from __future__ import annotations

import csv
import re
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "inputs" / "Especias_Benavidez_Precios.xlsx"
OUT = ROOT / "outputs"

LISTAS = [
    (1, "Lista 1", "Lista Base (Público)", 1.00),
    (2, "Lista 2", "Lista Minorista Sugerido", 1.15),
    (3, "Lista 3", "Lista Mayorista Especial", 0.90),
    (4, "Lista 4", "Lista Gran Distribuidor", 0.85),
]

# Rubros de alta rotación para carnicería elaboradora (buyer persona).
LEADER_KEYS = [
    "INTEGRAL DE MILANESA",
    "INTEGRAL MILANESA",
    "INTEGRAL DE HAMBURGUESA",
    "INTEGRAL HAMBURGUESA",
    "CHIMICHURRI",
    "REBOZADOR",
    "AJI MOLIDO",
    "AJÍ MOLIDO",
    "AJO DESHIDRATADO",
    "PIMIENTA",
    "TRIPAS",
    "MAD DE",
    "HUEVO EN POLVO",
    "SOJA",
    "CONDIMENTO PARA",
]

PACK_RE = re.compile(
    r"""(?:
        \(\s*B\s*/\s*(\d{1,3})\s*\) |
        \bPACK\s*(?:X\s*|DE\s*)?(\d{1,3})\b |
        \bX\s*(\d{1,3})\s*(?:U(?:NI(?:DADES?|D)?)?|UNID)\b |
        \bX\s*(\d{1,3})\s*$
    )""",
    re.I | re.VERBOSE,
)

SIZE_RE = re.compile(
    r"\b(\d+(?:[.,]\d+)?)\s*(KGS?|KILOS?|GRS?|GMS?|MG|MTS?|METROS?|LTS?|LITROS?|ML|CM|MM|UNIDADES?)\b",
    re.I,
)

_CAT1_RAW = {
    "ESPECIAS POR 1KG": ("Especias", "1 kg"),
    "ESPECIAS POR 500G": ("Especias", "500 g"),
    "ESPECIAS POR 250G": ("Especias", "250 g"),
    "ESPECIAS POR 5KG Y 2,5KG": ("Especias", "2.5–5 kg"),
    "INTEGRALES POR 1 KG": ("Integrales", "1 kg"),
    "INTEGRALES POR 5 KG": ("Integrales", "5 kg"),
    "INTEGRALES POR 10 KG": ("Integrales", "10 kg"),
    "INTEGRALES POR 20 KG": ("Integrales", "20 kg"),
    "FRACCIONADOS - BOLSAS PEQUENAS": ("Fraccionados", "Bolsas chicas"),
    "ADITIVOS": ("Aditivos", "Aditivos"),
    "ANTIOXIDANTES": ("Antioxidantes", "Antioxidantes"),
    "TRIPAS": ("Tripas", "Tripas"),
    "HILOS": ("Hilos", "Hilos de atar"),
    "CUCHILLOS": ("Cuchillos", "Cuchillos"),
    "CUCHILLAS Y GRILLAS": ("Cuchillas y grillas", "Cuchillas y grillas"),
    "MAQUINAS Y EQUIPAMIENTO": ("Máquinas", "Equipamiento"),
    "RESPUESTOS Y ACCESORIOS": ("Repuestos", "Accesorios"),
    "RESPUESTOS Y ACCESORIOS - BANDEJAS -": ("Repuestos", "Bandejas"),
    "ROPA DE FRIGORIFICO": ("Indumentaria", "Ropa de frigorífico"),
    "BOLSAS": ("Bolsas", "Bolsas"),
    "REPOSTERIA (SOLICITAR CON 1 SEMANA DE ANTICIPACION)": ("Repostería", "Pedido con anticipación"),
    "LEGUMBRES Y CEREALES (SOLICITAR CON 1 SEMANA DE ANTICIPACION)": ("Legumbres y cereales", "Pedido con anticipación"),
    "FRUTAS SECAS (SOLICITAR CON 1 SEMANA DE ANTICIPACION)": ("Frutas secas", "Pedido con anticipación"),
    "FRUTOS SECOS (SOLICITAR CON 1 SEMANA DE ANTICIPACION)": ("Frutos secos", "Pedido con anticipación"),
    "SEMILLAS (SOLICITAR CON 1 SEMANA DE ANTICIPACION)": ("Semillas", "Pedido con anticipación"),
}

NOUNS = [
    (["INTEGRAL"], "Integral para elaborados"),
    (["TRIPAS", "MAD DE", "BOWEL"], "Tripa para embutidos"),
    (["CUCHILLO"], "Cuchillo"),
    (["CUCHILLA", "GRILLA"], "Cuchilla o grilla"),
    (["EMBUTIDORA", "CLIPEADORA", "MOLEDORA", "SIERRA", "AMASADORA"], "Máquina"),
    (["BANDEJA"], "Bandeja"),
    (["RESPUESTO", "ESPIRAL", "PIÑON", "ENGRANAJE", "CORREA"], "Repuesto"),
    (["HILO", "BOBINA"], "Hilo para atar"),
    (["DELANTAL", "BOTA", "GUANTE", "COFIA", "CHAQUETA", "PANTALON", "CAPA"], "Indumentaria de frigorífico"),
    (["BOLSA"], "Bolsa"),
    (["ADITIVO", "AGLUSOR", "REBOZADOR", "HUEVO EN POLVO"], "Aditivo"),
    (["ANTIOXIDANTE", "AGUA OXIGENADA"], "Antioxidante"),
    (["FRACCIONADO"], "Fraccionado de condimento"),
    (["SEMILLA"], "Semilla"),
    (["LEGUMBRE", "POROTO", "LENTEJA", "GARBANZO", "ARROZ", "CEREAL"], "Legumbre o cereal"),
    (["NUEZ", "ALMENDRA", "PASAS", "FRUTO SECO", "FRUTA SECA"], "Fruto seco"),
    (["HARINA", "AZUCAR", "REPOSTER"], "Insumo de repostería"),
]

IMAGES = {
    "Especias": "https://images.unsplash.com/photo-1596040033229-a9821ebd058d?w=400",
    "Integrales": "https://images.unsplash.com/photo-1596040033229-a9821ebd058d?w=400",
    "Fraccionados": "https://images.unsplash.com/photo-1596040033229-a9821ebd058d?w=400",
    "Aditivos": "https://images.unsplash.com/photo-1582719471384-894fbb16e074?w=400",
    "Antioxidantes": "https://images.unsplash.com/photo-1582719471384-894fbb16e074?w=400",
    "Tripas": "https://images.unsplash.com/photo-1607623814075-e51df1bdc66d?w=400",
    "Hilos": "https://images.unsplash.com/photo-1607623814075-e51df1bdc66d?w=400",
    "Cuchillos": "https://images.unsplash.com/photo-1593618998160-e34014e67546?w=400",
    "Cuchillas y grillas": "https://images.unsplash.com/photo-1593618998160-e34014e67546?w=400",
    "Máquinas": "https://images.unsplash.com/photo-1581091226825-a6a2a5aee158?w=400",
    "Repuestos": "https://images.unsplash.com/photo-1581091226825-a6a2a5aee158?w=400",
    "Indumentaria": "https://images.unsplash.com/photo-1581091226825-a6a2a5aee158?w=400",
    "Bolsas": "https://images.unsplash.com/photo-1610484826967-09c5720778c7?w=400",
    "Repostería": "https://images.unsplash.com/photo-1486427944299-d1955d23e34d?w=400",
    "Legumbres y cereales": "https://images.unsplash.com/photo-1515543904379-3d757afe72e4?w=400",
    "Frutas secas": "https://images.unsplash.com/photo-1599599810769-bcde5a160d25?w=400",
    "Frutos secos": "https://images.unsplash.com/photo-1599599810769-bcde5a160d25?w=400",
    "Semillas": "https://images.unsplash.com/photo-1515543904379-3d757afe72e4?w=400",
}

BRANDS = [
    "ESPECIAS BENAVIDEZ",
    "RUEDO",
    "TRINIDAD",
    "FREIRE",
    "SOL-",
    "SOL ",
]


def fold(s: str) -> str:
    n = unicodedata.normalize("NFKD", s or "")
    return "".join(c for c in n if not unicodedata.combining(c)).upper()


CAT1_MAP = {fold(k): v for k, v in _CAT1_RAW.items()}


def clean_name(raw: str) -> str:
    n = re.sub(r"\s+", " ", str(raw or "").strip())
    return re.sub(r"\s*[-–.]+$", "", n).strip()


def detect_brand(name_up: str) -> str:
    for b in BRANDS:
        if b in name_up:
            if b.startswith("SOL"):
                return "Sol"
            return b.title() if b != "ESPECIAS BENAVIDEZ" else "Especias Benavidez"
    return "Especias Benavidez"


def noun_for(name_up: str, cat1: str) -> str:
    for keys, noun in NOUNS:
        if any(k in name_up for k in keys):
            return noun
    defaults = {
        "Especias": "Especia",
        "Integrales": "Integral para elaborados",
        "Fraccionados": "Fraccionado de condimento",
        "Aditivos": "Aditivo",
        "Antioxidantes": "Antioxidante",
        "Tripas": "Tripa para embutidos",
        "Hilos": "Hilo para atar",
        "Cuchillos": "Cuchillo",
        "Cuchillas y grillas": "Cuchilla o grilla",
        "Máquinas": "Máquina",
        "Repuestos": "Repuesto",
        "Indumentaria": "Indumentaria de frigorífico",
        "Bolsas": "Bolsa",
        "Repostería": "Insumo de repostería",
        "Legumbres y cereales": "Legumbre o cereal",
        "Frutas secas": "Fruta seca",
        "Frutos secos": "Fruto seco",
        "Semillas": "Semilla",
    }
    return defaults.get(cat1, "Producto")


def categoria_3(name_up: str, cat1: str) -> str:
    families = [
        ("MILANESA", "Integrales milanesas"),
        ("HAMBURGUESA", "Integrales hamburguesas"),
        ("CHORIZO", "Integrales chorizo"),
        ("MORCILLA", "Integrales morcilla"),
        ("SALAME", "Integrales salame"),
        ("CHIMICHURRI", "Chimichurri"),
        ("AJI", "Ají"),
        ("AJO", "Ajo"),
        ("PIMIENTA", "Pimienta"),
        ("CANELA", "Canela"),
        ("COMINO", "Comino"),
        ("OREGANO", "Orégano"),
        ("PIMENTON", "Pimentón"),
        ("CURRY", "Curry"),
        ("NUEZ MOSCADA", "Nuez moscada"),
        ("CLAVO", "Clavo de olor"),
        ("ANIS", "Anís"),
        ("AZAFRAN", "Azafrán"),
        ("CEBOLLA", "Cebolla deshidratada"),
        ("PEREJIL", "Perejil"),
        ("ALBAHACA", "Albahaca"),
        ("REBOZADOR", "Rebozador"),
        ("EMBUTIDORA", "Embutidoras"),
        ("CLIPEADORA", "Clipeadoras"),
        ("MOLEDORA", "Moledoras"),
        ("SIERRA", "Sierras"),
        ("CUCHILLO", "Cuchillos"),
        ("CUCHILLA", "Cuchillas"),
        ("GRILLA", "Grillas"),
        ("BANDEJA", "Bandejas"),
        ("ESPIRAL", "Espirales"),
        ("HILO", "Hilos"),
        ("DELANTAL", "Delantales"),
        ("BOTA", "Botas"),
        ("GUANTE", "Guantes"),
    ]
    for key, fam in families:
        if key in name_up:
            return fam
    return cat1


def unidades_por_bulto(name: str) -> int:
    if re.search(r"X\s*UNIDAD\b", name, re.I):
        return 1
    m = PACK_RE.search(name)
    if not m:
        return 1
    for g in m.groups():
        if g:
            n = int(g)
            return n if 1 < n <= 200 else 1
    return 1


def extract_size(name: str) -> str:
    m = SIZE_RE.search(name)
    if not m:
        return ""
    qty = m.group(1).replace(",", ".")
    unit = m.group(2).lower()
    unit_map = {
        "kg": "kg",
        "kgs": "kg",
        "kilo": "kg",
        "kilos": "kg",
        "gr": "g",
        "grs": "g",
        "gms": "g",
        "mg": "mg",
        "mt": "m",
        "mts": "m",
        "metro": "m",
        "metros": "m",
        "lt": "L",
        "lts": "L",
        "litro": "L",
        "litros": "L",
        "ml": "ml",
        "cm": "cm",
        "mm": "mm",
        "unidad": "un.",
        "unidades": "un.",
    }
    return f"{qty} {unit_map.get(unit, unit)}"


def aliases(nombre: str, brand: str, code: str, cat3: str) -> str:
    parts = [nombre, code, cat3]
    if brand:
        parts.append(brand)
    words = [w for w in re.split(r"[\s\-]+", nombre) if w]
    if len(words) >= 2:
        parts.append(" ".join(words[:3]))
    compact = re.sub(r"\s+", " ", re.sub(r"[^A-Za-z0-9ÁÉÍÓÚÑáéíóúñ ]+", " ", nombre)).strip()
    if compact and compact != nombre:
        parts.append(compact)
    seen: set[str] = set()
    out: list[str] = []
    for p in parts:
        k = fold(p)
        if k and k not in seen:
            seen.add(k)
            out.append(p)
    return "|".join(out[:6])


def pretty_name(nombre: str) -> str:
    titled = nombre.title().replace(" De ", " de ").replace(" En ", " en ").replace(" Para ", " para ")
    titled = titled.replace("X ", "x ").replace("Kg", "kg").replace("Gr", "g")
    return titled


def descripcion(noun: str, nombre: str, brand: str, size: str, bulto: int, cat2: str) -> str:
    bits = [f"{noun} {pretty_name(nombre)}"]
    bits.append(f"marca {brand}")
    if size:
        bits.append(size)
    elif cat2 and cat2 not in {"Aditivos", "Antioxidantes", "Tripas"}:
        bits.append(f"presentación {cat2}")
    if bulto > 1:
        bits.append(f"x{bulto} unidades por bulto")
    text = ", ".join(bits) + "."
    words = text.split()
    if len(words) > 25:
        text = " ".join(words[:24]).rstrip(",.") + "."
    if len(text.split()) < 10:
        text = text.rstrip(".") + ", línea Especias Benavidez Córdoba."
    return text


def is_leader(name_up: str, cat1: str) -> bool:
    if any(k in name_up for k in LEADER_KEYS):
        return True
    return cat1 in {"Especias", "Integrales", "Tripas", "Fraccionados"} and any(
        k in name_up for k in ("AJI", "AJO", "PIMIENTA", "CHIMICHURRI", "INTEGRAL", "MAD ", "TRIPAS")
    )


def main() -> int:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Precios"]
    raw_rows = list(ws.iter_rows(min_row=2, values_only=True))

    seen_codes: Counter[str] = Counter()
    accepted: list[dict] = []
    rejected: list[dict] = []

    for excel_row, (code_raw, name_raw, price, cat_raw) in enumerate(raw_rows, start=2):
        code = str(code_raw or "").strip()
        nombre = clean_name(name_raw)
        cat_excel = str(cat_raw or "").strip()
        price_f = float(price or 0)
        reason = ""
        if not nombre:
            reason = "nombre vacío"
        elif not code:
            reason = "SKU vacío"
        elif price_f <= 0:
            reason = "precio <= 0 (consultar / sin lista)"
        if reason:
            rejected.append(
                {
                    "excel_row": excel_row,
                    "product_code": code,
                    "nombre": nombre or str(name_raw or ""),
                    "precio": price_f,
                    "categoria": cat_excel,
                    "motivo": reason,
                }
            )
            continue
        seen_codes[code] += 1
        if seen_codes[code] > 1:
            code = f"{code}-{seen_codes[code]}"
        cat1, cat2 = CAT1_MAP.get(fold(cat_excel), (cat_excel.title() or "General", "General"))
        accepted.append(
            {
                "product_code": code,
                "nombre": nombre,
                "precio": price_f,
                "cat_excel": cat_excel,
                "cat1": cat1,
                "cat2": cat2,
            }
        )

    scored = []
    for i, row in enumerate(accepted):
        up = fold(row["nombre"])
        leader = is_leader(up, row["cat1"])
        scored.append((0 if leader else 1, i, row))
    scored.sort()

    n = len(scored)
    cut = max(1, int(n * 0.20))
    products = []
    for rank, (_g, _i, row) in enumerate(scored):
        nombre = row["nombre"]
        up = fold(nombre)
        brand = detect_brand(up)
        cat1, cat2 = row["cat1"], row["cat2"]
        cat3 = categoria_3(up, cat1)
        cat4 = brand
        bulto = unidades_por_bulto(nombre)
        size = extract_size(nombre)
        noun = noun_for(up, cat1)
        if rank < cut:
            rot = round(0.95 - (rank / cut) * 0.20, 4)
            prio = round(1.0 - (rank / cut) * 0.50, 4)
        else:
            span = max(1, n - cut)
            rot = round(0.70 - ((rank - cut) / span) * 0.60, 4)
            prio = round(0.45 - ((rank - cut) / span) * 0.40, 4)
        rot = max(0.10, min(0.95, rot))
        prio = max(0.05, min(1.0, prio))
        # Repuestos y máquinas: bajar un poco la cola
        if cat1 in {"Repuestos", "Máquinas"}:
            rot = max(0.10, round(rot * 0.75, 4))
            prio = max(0.05, round(prio * 0.70, 4))
        stock = int(10 + rot * 490)
        products.append(
            {
                "product_code": row["product_code"],
                "nombre": nombre,
                "precio_lista_1": f"{row['precio']:.2f}",
                "stock": stock,
                "unidades_por_bulto": bulto,
                "unidad_minima_de_venta": "unidad",
                "umv_tipo": "unidad",
                "categoria_1": cat1,
                "categoria_2": cat2,
                "categoria_3": cat3,
                "categoria_4": cat4,
                "aliases": aliases(nombre, brand, row["product_code"], cat3),
                "rotacion_index": f"{rot:.4f}",
                "mental_priority": f"{prio:.4f}",
                "descripcion": descripcion(noun, nombre, brand, size, bulto, cat2),
                "image_url": IMAGES.get(cat1, IMAGES["Especias"]),
                "en_catalogo": "true",
                "is_mock": "true",
                "fuente_hoja": "Precios",
                "_precio": row["precio"],
                "_brand": brand,
            }
        )

    prod_fields = [
        "product_code",
        "nombre",
        "precio_lista_1",
        "stock",
        "unidades_por_bulto",
        "unidad_minima_de_venta",
        "umv_tipo",
        "categoria_1",
        "categoria_2",
        "categoria_3",
        "categoria_4",
        "aliases",
        "rotacion_index",
        "mental_priority",
        "descripcion",
        "image_url",
        "en_catalogo",
        "is_mock",
        "fuente_hoja",
    ]
    with (OUT / "phase-01-productos.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=prod_fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(products)

    for lid, _nom, _desc, mult in LISTAS:
        path = OUT / f"phase-01-lista-precios-{lid}.csv"
        with path.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["product_code", "precio_unidad", "is_mock"])
            w.writeheader()
            for p in products:
                w.writerow(
                    {
                        "product_code": p["product_code"],
                        "precio_unidad": f"{round(p['_precio'] * mult, 2):.2f}",
                        "is_mock": "true",
                    }
                )

    with (OUT / "phase-01-listas-precios.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["lista_precios_id", "nombre", "descripcion", "multiplicador_sobre_lista_1", "is_mock"],
        )
        w.writeheader()
        for lid, nom, desc, mult in LISTAS:
            w.writerow(
                {
                    "lista_precios_id": lid,
                    "nombre": nom,
                    "descripcion": desc,
                    "multiplicador_sobre_lista_1": f"{mult:.2f}",
                    "is_mock": "true",
                }
            )

    with (OUT / "phase-01-rechazados.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(
            f, fieldnames=["excel_row", "product_code", "nombre", "precio", "categoria", "motivo"]
        )
        w.writeheader()
        w.writerows(rejected)

    cats = Counter(p["categoria_1"] for p in products)
    brands = Counter(p["_brand"] for p in products)
    codes = [p["product_code"] for p in products]
    assert len(codes) == len(set(codes))
    assert all(p["_precio"] > 0 for p in products)
    print(f"aceptados={len(products)} rechazados={len(rejected)}")
    print("categorias:", dict(cats.most_common()))
    print("marcas:", brands.most_common(6))
    print("sku_unicos=ok precio>0=ok")
    print("muestra:", products[0]["product_code"], products[0]["nombre"], products[0]["descripcion"])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
