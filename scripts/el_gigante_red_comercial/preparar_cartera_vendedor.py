#!/usr/bin/env python3
"""
Prepara cartera de un vendedor (Ivan / Fernando / …) → CSV + hulls Google.

  python scripts/el_gigante_red_comercial/preparar_cartera_vendedor.py --vendedor Ivan
  python scripts/el_gigante_red_comercial/preparar_cartera_vendedor.py --vendedor Fernando
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import time
import unicodedata
from pathlib import Path
from urllib.parse import quote

import httpx
import openpyxl
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parents[2]
load_dotenv(ROOT / ".env")
load_dotenv()

try:
    from shapely.geometry import MultiPoint, mapping
except ImportError as e:
    raise SystemExit("Instalá shapely: pip install shapely") from e

# filename → (nombre zona BD, dias_visita). Días siempre desde filename.
VENDOR_CONFIG: dict[str, dict] = {
    "Ivan": {
        "input_dir": "ivan_ruta_2",
        "slug": "ivan",
        "files": {
            "LUNES Y JUEVES - B° VIRGEN DEL ROSARIO.xlsx": ("Virgen del Rosario", ["lunes", "jueves"]),
            "LUNES Y JUEVES - SAN JOSE OBRERO.xlsx": ("San jose obrero", ["lunes", "jueves"]),
            "MARTES - LA PILAR.xlsx": ("La pilar", ["martes"]),
            "MARTES - OBRERO.xlsx": ("Obrero", ["martes"]),
            "MIERCOLES Y VIERNES - SAN JUAN BAUTISTA.xlsx": ("San Juan Bautista", ["miércoles", "viernes"]),
        },
    },
    "Fernando": {
        "input_dir": "fernando_ruta_3",
        "slug": "fernando",
        "files": {
            "LUNES Y JUEVES - LOTE 4.xlsx": ("Lote 4", ["lunes", "jueves"]),
            "LUNES Y JUEVES - VILLA HERMOSA.xlsx": ("Villa hermosa", ["lunes", "jueves"]),
            "MIERCOLES Y VIERNES - VIAL.xlsx": ("B° vial", ["miércoles", "viernes"]),
            "MIERCOLES Y VIERNES - VILLA LOURDES.xlsx": ("Villa Lourdes", ["miércoles", "viernes"]),
        },
    },
    # Workbooks multi-hoja: cada sheet = una zona (nombre canónico BD).
    "Javier": {
        "input_dir": "javier_ruta_4",
        "slug": "javier",
        "files": {
            "LUNES Y JUEVES - FONTANA - GUADALUPE - TERMINAL - PARQUE URBANO.xlsx": {
                "dias": ["lunes", "jueves"],
                "sheets": {
                    "FONTANA": "Fontana",
                    "GUADALUPE": "Guadalupe",
                    "TERMINAL": "Terminal",
                    "PARQUE URBANO": "Parque Urbano",
                },
            },
            "MARTES - CENTRO 2.xlsx": {
                "dias": ["martes"],
                "sheets": {"CENTRO 2": "Centro 2"},
            },
            "MIERCOLES Y VIERNES - MARIANO MORENO - DON BOSCO.xlsx": {
                "dias": ["miércoles", "viernes"],
                "sheets": {
                    "MARIANO MORENO": "Mariano Moreno",
                    "DON BOSCO": "Don Bosco",
                },
            },
        },
    },
    "Enrique": {
        "input_dir": "enrique_ruta_5",
        "slug": "enrique",
        "files": {
            "LUNES Y MIERCOLES - LAGUNA SIAM - 7 DE NOVIEMBRE.xlsx": {
                "dias": ["lunes", "miércoles"],
                "sheets": {
                    "LAGUNA SIAM": "Laguna Siam",
                    "7 DE NOVIEMBRE": "7 de Noviembre",
                },
            },
            "MARTES - 2 DE ABRIL - JUAN MANUEL DE ROSAS.xlsx": {
                "dias": ["martes"],
                "sheets": {
                    "2 DE ABRIL": "2 de Abril",
                    "JUAN MANUEL DE ROSAS": "Juan Manuel de Rosas",
                },
            },
            "MIERCOLES Y VIERNES - SAN MIGUEL - SAN AGUSTIN.xlsx": {
                "dias": ["miércoles", "viernes"],
                "sheets": {
                    "SAN MIGUEL": "San Miguel",
                    "SAN AGUSTIN": "San Agustin",
                },
            },
        },
    },
}

ENTREGA_NEXT = {
    "lunes": "martes",
    "martes": "miércoles",
    "miércoles": "jueves",
    "jueves": "viernes",
    "viernes": "sábado",
    "sábado": "lunes",
    "domingo": "lunes",
}

_GOOGLE_DISABLED = False


def _slug(name: str) -> str:
    s = unicodedata.normalize("NFKD", name)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s.strip().lower()).strip("-")
    return s or "zona"


def parse_excel_coord(raw) -> float | None:
    if raw is None or raw == "":
        return None
    try:
        v = float(raw)
    except (TypeError, ValueError):
        return None
    if abs(v) > 180:
        s = str(int(v)) if abs(v - int(v)) < 1e-9 else str(v)
        s = s.lstrip("-")
        if len(s) >= 8:
            sign = -1 if float(raw) < 0 else 1
            return sign * float(s[:2] + "." + s[2:])
    if -90 <= v <= 90 or -180 <= v <= 180:
        return v
    return None


def normalize_domicilio(raw: str | None) -> str:
    if not raw:
        return ""
    s = str(raw).strip()
    if " / " in s:
        s = s.split(" / ", 1)[0].strip()
    return re.sub(r"\s+", " ", s)


def find_header_row(rows: list) -> int:
    for i, row in enumerate(rows[:5]):
        if not row:
            continue
        first = str(row[0] or "").strip().lower().rstrip(".")
        if first in ("cod", "código", "codigo"):
            return i
    return 1  # default legacy (meta + header)


async def geocode_google(client: httpx.AsyncClient, address: str, api_key: str):
    url = (
        "https://maps.googleapis.com/maps/api/geocode/json"
        f"?address={quote(address)}&key={api_key}&region=ar&language=es"
    )
    r = await client.get(url, timeout=30)
    data = r.json()
    status = str(data.get("status") or "")
    results = data.get("results") or []
    if status != "OK" or not results:
        return None, None, status or "failed"
    loc = results[0].get("geometry", {}).get("location") or {}
    lat, lng = loc.get("lat"), loc.get("lng")
    if lat is None or lng is None:
        return None, None, status
    return float(lat), float(lng), "resolved"


async def geocode_nominatim(client: httpx.AsyncClient, address: str):
    r = await client.get(
        "https://nominatim.openstreetmap.org/search",
        params={"q": address, "format": "json", "limit": 1, "countrycodes": "ar"},
        headers={"User-Agent": "suplai-el-gigante-geocode/1.0"},
        timeout=30,
    )
    if r.status_code != 200:
        return None, None, f"nominatim_http_{r.status_code}"
    results = r.json() or []
    if not results:
        return None, None, "nominatim_zero"
    return float(results[0]["lat"]), float(results[0]["lon"]), "nominatim"


async def geocode(client, address, api_key):
    global _GOOGLE_DISABLED
    if api_key and not _GOOGLE_DISABLED:
        lat, lng, status = await geocode_google(client, address, api_key)
        if lat is not None:
            return lat, lng, status
        if status in ("REQUEST_DENIED", "OVER_QUERY_LIMIT"):
            _GOOGLE_DISABLED = True
    return await geocode_nominatim(client, address)


def hull_geojson(points: list[tuple[float, float]], buffer_m: float = 300.0) -> dict:
    mp = MultiPoint([(lng, lat) for lat, lng in points])
    buf_deg = buffer_m / 111_000.0
    geom = mp.convex_hull.buffer(buf_deg)
    if geom.geom_type == "Polygon":
        gj = mapping(geom)
        return {"type": "MultiPolygon", "coordinates": [gj["coordinates"]]}
    if geom.geom_type == "MultiPolygon":
        return mapping(geom)
    minx, miny, maxx, maxy = mp.bounds
    ring = [
        [minx - buf_deg, miny - buf_deg],
        [maxx + buf_deg, miny - buf_deg],
        [maxx + buf_deg, maxy + buf_deg],
        [minx - buf_deg, maxy + buf_deg],
        [minx - buf_deg, miny - buf_deg],
    ]
    return {"type": "MultiPolygon", "coordinates": [[ring]]}


def geojson_to_wkt(gj: dict) -> str:
    polys = []
    for poly in gj["coordinates"]:
        rings = []
        for ring in poly:
            coords = ", ".join(f"{x} {y}" for x, y in ring)
            rings.append(f"({coords})")
        polys.append(f"({','.join(rings)})")
    return f"SRID=4326;MULTIPOLYGON({','.join(polys)})"


def in_formosa(lat: float, lng: float) -> bool:
    return -26.35 <= lat <= -26.05 and -58.35 <= lng <= -58.05


def _rows_to_clients(
    rows: list,
    zona: str,
    dias: list[str],
    vendedor: str,
    source_label: str,
) -> list[dict]:
    hdr_i = find_header_row(rows)
    primary = dias[0]
    entrega = ENTREGA_NEXT.get(primary, "martes")
    clients = []
    for row in rows[hdr_i + 1 :]:
        if not row or row[0] is None:
            continue
        if str(row[0]).strip().lower().startswith("cod"):
            continue
        lat = parse_excel_coord(row[8] if len(row) > 8 else None)
        lng = parse_excel_coord(row[9] if len(row) > 9 else None)
        doc = row[1] if len(row) > 1 else None
        cliente = row[2] if len(row) > 2 else None
        fantasia = row[3] if len(row) > 3 else None
        localidad = row[4] if len(row) > 4 else zona
        domicilio = normalize_domicilio(row[6] if len(row) > 6 else "")
        tel = row[7] if len(row) > 7 else None
        clients.append(
            {
                "codigo": str(row[0]).strip(),
                "cuit": str(doc).strip() if doc not in (None, 0, "0") else "",
                "razon_social": str(cliente or "").strip(),
                "nombre": str(fantasia or cliente or "").strip(),
                "localidad": str(localidad or zona).strip(),
                "domicilio": domicilio,
                "phone_number": re.sub(r"\D", "", str(tel)) if tel else "",
                "lat": lat,
                "lng": lng,
                "geocode_status": "excel" if lat is not None and lng is not None else "pending",
                "lista_precios_id": 1,
                "dia_de_visita": primary,
                "dia_de_entrega": entrega,
                "dias_zona": ",".join(dias),
                "vendedor_nombre": vendedor,
                "zona_nombre": zona,
                "source_file": source_label,
                "is_mock": "false",
            }
        )
    return clients


def load_clients_from_xlsx(
    xlsx: Path, zona: str, dias: list[str], vendedor: str
) -> list[dict]:
    """Legacy: primera hoja → una zona (Ivan/Fernando/Diego)."""
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    return _rows_to_clients(rows, zona, dias, vendedor, xlsx.name)


def load_clients_from_workbook_sheets(
    xlsx: Path,
    sheet_to_zona: dict[str, str],
    dias: list[str],
    vendedor: str,
) -> dict[str, list[dict]]:
    """Multi-hoja: {zona_bd: clients[]}."""
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    by_zona: dict[str, list[dict]] = {}
    for sheet_name, zona in sheet_to_zona.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ! sheet faltante '{sheet_name}' en {xlsx.name}")
            continue
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        by_zona[zona] = _rows_to_clients(
            rows, zona, dias, vendedor, f"{xlsx.name}#{sheet_name}"
        )
    wb.close()
    return by_zona


async def geocode_clients(clients: list[dict], api_key: str, barrio_hint: str) -> None:
    async with httpx.AsyncClient() as client:
        for c in clients:
            if c["lat"] is not None and c["lng"] is not None:
                continue
            address = f"{c['domicilio']}, {barrio_hint}, Formosa, Formosa, Argentina"
            lat, lng, status = await geocode(client, address, api_key)
            if lat is None and api_key and not _GOOGLE_DISABLED:
                lat, lng, status = await geocode_google(
                    client, f"{c['domicilio']}, Formosa, Formosa, Argentina", api_key
                )
                time.sleep(0.05)
            if lat is None:
                lat, lng, status = await geocode_nominatim(
                    client, f"{c['domicilio']}, Formosa, Argentina"
                )
                time.sleep(1.05)
            c["lat"], c["lng"] = lat, lng
            c["geocode_status"] = status
            time.sleep(0.08)


def write_zone_outputs(out: Path, slug_vendor: str, zona: str, clients: list[dict], patch_zonas: bool) -> None:
    zslug = _slug(zona)
    csv_path = out / f"phase-04-clientes-{slug_vendor}-{zslug}.csv"
    keys = list(clients[0].keys()) if clients else []
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        w.writerows(clients)

    points = [
        (float(c["lat"]), float(c["lng"]))
        for c in clients
        if c["lat"] is not None and c["lng"] is not None and in_formosa(float(c["lat"]), float(c["lng"]))
    ]
    ok = sum(1 for c in clients if c["lat"] is not None)
    print(f"  [{zona}] n={len(clients)} coords={ok} → {csv_path.name}")
    if not points:
        print(f"  [{zona}] hull omitido")
        return

    gj = hull_geojson(points, buffer_m=300 if len(points) >= 3 else 400)
    wkt = geojson_to_wkt(gj)
    (out / f"phase-04-zona-{zslug}-polygon.geojson").write_text(
        json.dumps(gj, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out / f"phase-04-zona-{zslug}-polygon.wkt").write_text(wkt, encoding="utf-8")

    if not patch_zonas:
        return
    zonas_csv = out / "phase-04-zonas-real.csv"
    if not zonas_csv.exists():
        return
    with zonas_csv.open(encoding="utf-8") as f:
        rows_z = list(csv.DictReader(f))
    dias = (clients[0].get("dias_zona") or "lunes").split(",")
    matched = False
    for z in rows_z:
        if z.get("nombre", "").strip().lower() == zona.strip().lower():
            z["geometry_wkt"] = wkt
            z["dias_visita"] = ",".join(dias)
            z["dia_visita"] = dias[0]
            matched = True
    if not matched:
        # Enrique / zonas nuevas: append fila
        vidx = {"Diego": "0", "Ivan": "1", "Fernando": "2", "Javier": "3", "Enrique": "4"}.get(
            clients[0].get("vendedor_nombre", ""), "4"
        )
        rows_z.append(
            {
                "nombre": zona,
                "zone_type": "route",
                "dias_visita": ",".join(dias),
                "dia_visita": dias[0],
                "codigo_ruta": f"RUTA-{_slug(zona).upper()[:12]}",
                "vendedor_idx": vidx,
                "geometry_wkt": wkt,
                "color": "#39FF14",
                "is_mock": "false",
            }
        )
        print(f"  [{zona}] append zonas-real.csv (nueva)")
    with zonas_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows_z[0].keys()))
        w.writeheader()
        w.writerows(rows_z)
    if matched:
        print(f"  [{zona}] patch zonas-real.csv")


async def main_async(esquema: str, vendedor: str, dry_run: bool) -> None:
    cfg = VENDOR_CONFIG[vendedor]
    base = ROOT / "implementacion" / esquema
    input_dir = base / "inputs" / cfg["input_dir"]
    out = base / "outputs"
    out.mkdir(parents=True, exist_ok=True)
    if not input_dir.is_dir():
        raise SystemExit(f"Falta {input_dir}")

    api_key = (os.getenv("GOOGLE_MAPS_API_KEY") or "").strip()
    seen_codigos: set[str] = set()
    all_clients: list[dict] = []
    skipped_dup = 0

    for fname, file_cfg in cfg["files"].items():
        xlsx = input_dir / fname
        if not xlsx.exists():
            print(f"  ! falta {fname}")
            continue

        # Compat: tuple (zona, dias) | dict {dias, sheets}
        if isinstance(file_cfg, tuple):
            zona, dias = file_cfg
            zones_payload = {zona: load_clients_from_xlsx(xlsx, zona, dias, vendedor)}
        else:
            dias = file_cfg["dias"]
            zones_payload = load_clients_from_workbook_sheets(
                xlsx, file_cfg["sheets"], dias, vendedor
            )

        for zona, clients in zones_payload.items():
            unique = []
            for c in clients:
                code = c["codigo"]
                if code in seen_codigos:
                    skipped_dup += 1
                    continue
                seen_codigos.add(code)
                unique.append(c)
            print(
                f"[*] {zona}: {len(unique)} filas (raw={len(clients)}) ← {fname} dias={dias}"
            )

            if dry_run:
                all_clients.extend(unique)
                continue

            need = sum(1 for c in unique if c["lat"] is None)
            if need:
                print(f"  geocoding {need}…")
                await geocode_clients(unique, api_key, barrio_hint=zona)
            if unique:
                write_zone_outputs(out, cfg["slug"], zona, unique, patch_zonas=True)
                all_clients.extend(unique)
            else:
                print(f"  [{zona}] sin clientes tras dedupe")

    cons = out / f"phase-04-clientes-{cfg['slug']}-all.csv"
    if all_clients:
        keys: list[str] = []
        for r in all_clients:
            for k in r:
                if k not in keys:
                    keys.append(k)
        with cons.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=keys)
            w.writeheader()
            w.writerows(all_clients)
    print(f"\nOK {vendedor}: {len(all_clients)} clientes → {cons.name} (dup omitidos={skipped_dup})")


def main() -> None:
    import asyncio

    ap = argparse.ArgumentParser()
    ap.add_argument("--esquema", default="el_gigante")
    ap.add_argument("--vendedor", required=True, choices=sorted(VENDOR_CONFIG.keys()))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    asyncio.run(main_async(args.esquema, args.vendedor, args.dry_run))


if __name__ == "__main__":
    main()
