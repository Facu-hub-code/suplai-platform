#!/usr/bin/env python3
"""Parsea rutas_vendedores.xlsx → CSVs de vendedores y zonas (con dias_visita multi-día)."""

from __future__ import annotations

import argparse
import csv
import math
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[2]
DIAS = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"]
DIA_CANON = {
    "lunes": "lunes",
    "martes": "martes",
    "miercoles": "miércoles",
    "miércoles": "miércoles",
    "jueves": "jueves",
    "viernes": "viernes",
    "sabado": "sábado",
    "sábado": "sábado",
    "domingo": "domingo",
}


def normalize_dia(raw: str | None) -> str | None:
    if not raw:
        return None
    key = str(raw).strip().lower()
    return DIA_CANON.get(key)


def normalize_zone_name(raw: str) -> str:
    s = re.sub(r"\s+", " ", str(raw).strip())
    return s


def placeholder_wkt(lat: float, lng: float, idx: int, radius_deg: float = 0.008) -> str:
    """Cuadrado pequeño alrededor del centro, desplazado por idx para no solapar."""
    # offset en grilla 3x3
    ox = (idx % 3) - 1
    oy = (idx // 3) - 1
    clat = lat + oy * radius_deg * 2.2
    clng = lng + ox * radius_deg * 2.2
    ring = [
        (clng - radius_deg, clat - radius_deg),
        (clng + radius_deg, clat - radius_deg),
        (clng + radius_deg, clat + radius_deg),
        (clng - radius_deg, clat + radius_deg),
        (clng - radius_deg, clat - radius_deg),
    ]
    coords = ", ".join(f"{x} {y}" for x, y in ring)
    return f"SRID=4326;MULTIPOLYGON((({coords})))"


def parse_rutas(xlsx: Path) -> tuple[list[dict], list[dict]]:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    vendedores: list[dict] = []
    # zone_key -> {nombre, vendedor_idx, dias:set, codigo_ruta}
    zones: dict[tuple[int, str], dict] = {}

    i = 0
    while i < len(rows):
        row = rows[i]
        if not row or row[0] is None:
            i += 1
            continue
        nombre = str(row[0]).strip()
        # header vendedor: name + "ruta N"
        ruta_cell = row[1]
        if ruta_cell is None or not re.search(r"ruta", str(ruta_cell), re.I):
            i += 1
            continue

        codigo_ruta = re.sub(r"\s+", " ", str(ruta_cell).strip())
        vendedor_idx = len(vendedores)
        vendedores.append(
            {
                "nombre": nombre.rstrip(),
                "telefono": f"5493716{100000 + vendedor_idx:06d}"[-13:],
                "email": f"{nombre.strip().lower().replace(' ', '.')}.el_gigante@suplaisales.placeholder",
                "zona": None,
                "codigo_ruta": codigo_ruta.upper().replace(" ", ""),
                "is_mock": "false",
            }
        )

        # next row = days header
        i += 1
        if i >= len(rows):
            break
        day_row = rows[i]
        dias_cols: list[str | None] = []
        for cell in day_row[:5]:
            dias_cols.append(normalize_dia(str(cell) if cell is not None else None))
        i += 1

        # zone rows until blank block
        while i < len(rows):
            zrow = rows[i]
            if zrow is None or all(c is None or str(c).strip() == "" for c in (zrow[:5] if zrow else [])):
                # peek if next is new vendor
                i += 1
                break
            # if looks like new vendor header
            if zrow[0] and zrow[1] and re.search(r"ruta", str(zrow[1]), re.I):
                break
            for col, dia in enumerate(dias_cols):
                if dia is None or col >= len(zrow):
                    continue
                cell = zrow[col]
                if cell is None or str(cell).strip() == "":
                    continue
                # may contain multiple zones separated by -
                parts = [normalize_zone_name(p) for p in re.split(r"\s*-\s*", str(cell)) if p.strip()]
                for zname in parts:
                    key = (vendedor_idx, zname.lower())
                    if key not in zones:
                        zones[key] = {
                            "nombre": zname,
                            "vendedor_idx": vendedor_idx,
                            "dias": set(),
                            "codigo_ruta": f"{vendedores[vendedor_idx]['codigo_ruta']}-{zname[:8].upper().replace(' ', '')}",
                        }
                    zones[key]["dias"].add(dia)
            i += 1

    # order dias by week
    order = {d: n for n, d in enumerate(["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"])}
    zona_rows = []
    for idx, z in enumerate(sorted(zones.values(), key=lambda x: (x["vendedor_idx"], x["nombre"]))):
        dias_sorted = sorted(z["dias"], key=lambda d: order.get(d, 99))
        zona_rows.append(
            {
                "nombre": z["nombre"],
                "zone_type": "route",
                "dias_visita": ",".join(dias_sorted),
                "dia_visita": dias_sorted[0] if dias_sorted else "lunes",
                "codigo_ruta": z["codigo_ruta"][:40],
                "vendedor_idx": z["vendedor_idx"],
                "geometry_wkt": "",  # filled later / placeholder
                "is_mock": "false",
                "color": f"#{(40 + idx * 37) % 200:02x}{(120 + idx * 53) % 200:02x}{(80 + idx * 19) % 200:02x}",
            }
        )
    return vendedores, zona_rows


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--esquema", default="el_gigante")
    args = ap.parse_args()
    esquema = args.esquema
    base = ROOT / "implementacion" / esquema
    manifest = yaml.safe_load((base / "manifest.yaml").read_text(encoding="utf-8"))
    centro = manifest.get("coordenadas_centro") or [-26.1775, -58.1781]
    lat, lng = float(centro[0]), float(centro[1])

    xlsx = base / "inputs" / "rutas_vendedores.xlsx"
    vendedores, zonas = parse_rutas(xlsx)
    for i, z in enumerate(zonas):
        z["geometry_wkt"] = placeholder_wkt(lat, lng, i)

    out = base / "outputs"
    out.mkdir(parents=True, exist_ok=True)
    vpath = out / "phase-04-vendedores-real.csv"
    zpath = out / "phase-04-zonas-real.csv"

    with vpath.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["nombre", "telefono", "email", "zona", "codigo_ruta", "is_mock"],
        )
        w.writeheader()
        w.writerows(vendedores)

    with zpath.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "nombre",
                "zone_type",
                "dias_visita",
                "dia_visita",
                "codigo_ruta",
                "vendedor_idx",
                "geometry_wkt",
                "color",
                "is_mock",
            ],
        )
        w.writeheader()
        w.writerows(zonas)

    print(f"OK vendedores={len(vendedores)} → {vpath}")
    print(f"OK zonas={len(zonas)} → {zpath}")
    for v in vendedores:
        print(f"  - {v['nombre']} ({v['codigo_ruta']})")
    for z in zonas:
        print(f"  - [{z['vendedor_idx']}] {z['nombre']} dias={z['dias_visita']}")


if __name__ == "__main__":
    main()
