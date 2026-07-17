#!/usr/bin/env python3
"""
Prepara carteras de Diego (Ruta 1) desde Excels → CSV clientes + polígonos por zona.

Uso:
  # Solo auditar / parsear (sin Google):
  python scripts/el_gigante_red_comercial/preparar_cartera_diego.py --esquema el_gigante --dry-run

  # Geocode + hull (requiere GOOGLE_MAPS_API_KEY server-side):
  python scripts/el_gigante_red_comercial/preparar_cartera_diego.py --esquema el_gigante

  # Una zona:
  python scripts/el_gigante_red_comercial/preparar_cartera_diego.py --zona "Centro 1"
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import time
import unicodedata
from pathlib import Path
from urllib.parse import quote

import httpx
import openpyxl
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parents[2]
load_dotenv(ROOT / ".env")
load_dotenv()

try:
    from shapely.geometry import MultiPoint, mapping
except ImportError as e:
    raise SystemExit("Instalá shapely: pip install shapely") from e

# Nombre canónico (como en phase-04-zonas-real / BD) + días.
# Preferimos filename; si meta discrepa, se reporta.
ZONE_FROM_FILENAME: dict[str, tuple[str, list[str]]] = {
    "LUNES Y JUEVES - SAN FRANCISCO.xlsx": ("San Francisco", ["lunes", "jueves"]),
    "LUNES_Y_JUEVES_SAN_FRANCISCO.xlsx": ("San Francisco", ["lunes", "jueves"]),
    "MARTES - CENTRO 1.xlsx": ("Centro 1", ["martes"]),
    "MARTES - SAN PEDRO.xlsx": ("San Pedro", ["martes"]),
    "MIERCOLES Y VIERNES -EVITA.xlsx": ("Evita", ["miércoles", "viernes"]),
    "MIERCOLES Y VIERNES -INCONE.xlsx": ("Incone", ["miércoles", "viernes"]),
    "MIERCOLES Y VIERNES -MALVINAS.xlsx": ("Malvinas", ["miércoles", "viernes"]),
    "MIERCOLES Y VIERNES -SAN ANDRES 1.xlsx": ("San Andres", ["miércoles", "viernes"]),
    "MIERCOLES Y VIERNES COLUCCIO.xlsx": ("Coluccio", ["miércoles", "viernes"]),
}

ENTREGA_NEXT = {
    "lunes": "martes",
    "martes": "miércoles",
    "miércoles": "jueves",
    "jueves": "viernes",
    "viernes": "sábado",
    "sábado": "lunes",
    "domingo": "lunes",
}

_GOOGLE_DISABLED = False


def _slug(name: str) -> str:
    s = unicodedata.normalize("NFKD", name)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s.strip().lower()).strip("-")
    return s or "zona"


def parse_excel_coord(raw) -> float | None:
    if raw is None or raw == "":
        return None
    try:
        v = float(raw)
    except (TypeError, ValueError):
        return None
    if abs(v) > 180:
        s = str(int(v)) if abs(v - int(v)) < 1e-9 else str(v)
        s = s.lstrip("-")
        if len(s) >= 8:
            sign = -1 if float(raw) < 0 else 1
            return sign * float(s[:2] + "." + s[2:])
    if -90 <= v <= 90 or -180 <= v <= 180:
        return v
    return None


def normalize_domicilio(raw: str | None) -> str:
    if not raw:
        return ""
    s = str(raw).strip()
    if " / " in s:
        s = s.split(" / ", 1)[0].strip()
    return re.sub(r"\s+", " ", s)


def parse_dias_from_meta(raw: str | None) -> list[str]:
    if not raw:
        return []
    s = unicodedata.normalize("NFKD", str(raw).lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.replace(" y ", " ").replace("-", " ").replace(",", " ")
    tokens = [t for t in s.split() if t]
    mapa = {
        "lunes": "lunes",
        "martes": "martes",
        "miercoles": "miércoles",
        "jueves": "jueves",
        "viernes": "viernes",
        "sabado": "sábado",
        "domingo": "domingo",
    }
    out = []
    for t in tokens:
        if t in mapa and mapa[t] not in out:
            out.append(mapa[t])
    return out


async def geocode_google(
    client: httpx.AsyncClient, address: str, api_key: str
) -> tuple[float | None, float | None, str]:
    url = (
        "https://maps.googleapis.com/maps/api/geocode/json"
        f"?address={quote(address)}&key={api_key}&region=ar&language=es"
    )
    r = await client.get(url, timeout=30)
    data = r.json()
    status = str(data.get("status") or "")
    results = data.get("results") or []
    if status != "OK" or not results:
        return None, None, status or "failed"
    loc = results[0].get("geometry", {}).get("location") or {}
    lat, lng = loc.get("lat"), loc.get("lng")
    if lat is None or lng is None:
        return None, None, status
    return float(lat), float(lng), "resolved"


async def geocode_nominatim(
    client: httpx.AsyncClient, address: str
) -> tuple[float | None, float | None, str]:
    r = await client.get(
        "https://nominatim.openstreetmap.org/search",
        params={"q": address, "format": "json", "limit": 1, "countrycodes": "ar"},
        headers={"User-Agent": "suplai-el-gigante-geocode/1.0 (implementacion)"},
        timeout=30,
    )
    if r.status_code != 200:
        return None, None, f"nominatim_http_{r.status_code}"
    results = r.json() or []
    if not results:
        return None, None, "nominatim_zero"
    lat, lng = results[0].get("lat"), results[0].get("lon")
    if lat is None or lng is None:
        return None, None, "nominatim_zero"
    return float(lat), float(lng), "nominatim"


async def geocode(
    client: httpx.AsyncClient, address: str, api_key: str
) -> tuple[float | None, float | None, str]:
    global _GOOGLE_DISABLED
    if api_key and not _GOOGLE_DISABLED:
        lat, lng, status = await geocode_google(client, address, api_key)
        if lat is not None and lng is not None:
            return lat, lng, status
        if status in ("REQUEST_DENIED", "OVER_QUERY_LIMIT"):
            _GOOGLE_DISABLED = True
    lat, lng, status = await geocode_nominatim(client, address)
    return lat, lng, status


def hull_geojson(points: list[tuple[float, float]], buffer_m: float = 300.0) -> dict:
    mp = MultiPoint([(lng, lat) for lat, lng in points])
    buf_deg = buffer_m / 111_000.0
    geom = mp.convex_hull.buffer(buf_deg)
    if geom.geom_type == "Polygon":
        gj = mapping(geom)
        return {"type": "MultiPolygon", "coordinates": [gj["coordinates"]]}
    if geom.geom_type == "MultiPolygon":
        return mapping(geom)
    minx, miny, maxx, maxy = mp.bounds
    ring = [
        [minx - buf_deg, miny - buf_deg],
        [maxx + buf_deg, miny - buf_deg],
        [maxx + buf_deg, maxy + buf_deg],
        [minx - buf_deg, maxy + buf_deg],
        [minx - buf_deg, miny - buf_deg],
    ]
    return {"type": "MultiPolygon", "coordinates": [[ring]]}


def geojson_to_wkt(gj: dict) -> str:
    assert gj["type"] == "MultiPolygon"
    polys = []
    for poly in gj["coordinates"]:
        rings = []
        for ring in poly:
            coords = ", ".join(f"{x} {y}" for x, y in ring)
            rings.append(f"({coords})")
        polys.append(f"({','.join(rings)})")
    return f"SRID=4326;MULTIPOLYGON({','.join(polys)})"


def in_formosa(lat: float, lng: float) -> bool:
    return -26.35 <= lat <= -26.05 and -58.35 <= lng <= -58.05


def discover_excels(input_dir: Path) -> list[Path]:
    files = sorted(input_dir.glob("*.xlsx"))
    # Preferir nombres con espacios (origen Desktop); evitar duplicar SF con underscore
    by_zone: dict[str, Path] = {}
    for f in files:
        mapped = ZONE_FROM_FILENAME.get(f.name)
        if not mapped:
            print(f"  ! Ignorado (sin mapeo): {f.name}")
            continue
        zone = mapped[0]
        # Preferir el path con espacios si hay dos
        prev = by_zone.get(zone)
        if prev is None or (" " in f.name and " " not in prev.name):
            by_zone[zone] = f
    return list(by_zone.values())


def load_clients_from_xlsx(
    xlsx: Path, zona_nombre: str, dias: list[str], vendedor: str = "Diego"
) -> tuple[list[dict], list[str]]:
    warnings: list[str] = []
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return [], [f"{xlsx.name}: vacío"]

    meta = rows[0]
    meta_dias = parse_dias_from_meta(str(meta[3]) if meta and len(meta) > 3 else "")
    if meta_dias and meta_dias != dias:
        warnings.append(
            f"{xlsx.name}: meta días={meta_dias} ≠ filename/mapeo={dias} → uso mapeo filename"
        )

    primary = dias[0]
    entrega = ENTREGA_NEXT.get(primary, "martes")
    clients: list[dict] = []
    for i, row in enumerate(rows):
        if i < 2 or not row or row[0] is None:
            continue
        codigo = row[0]
        doc = row[1]
        cliente = row[2]
        fantasia = row[3]
        localidad = row[4] or zona_nombre
        domicilio = normalize_domicilio(row[6])
        tel = row[7]
        lat = parse_excel_coord(row[8])
        lng = parse_excel_coord(row[9])
        clients.append(
            {
                "codigo": str(codigo).strip(),
                "cuit": str(doc).strip() if doc not in (None, 0, "0") else "",
                "razon_social": str(cliente or "").strip(),
                "nombre": str(fantasia or cliente or "").strip(),
                "localidad": str(localidad).strip(),
                "domicilio": domicilio,
                "phone_number": re.sub(r"\D", "", str(tel)) if tel else "",
                "lat": lat,
                "lng": lng,
                "geocode_status": "excel" if lat is not None and lng is not None else "pending",
                "lista_precios_id": 1,
                "dia_de_visita": primary,
                "dia_de_entrega": entrega,
                "dias_zona": ",".join(dias),
                "vendedor_nombre": vendedor,
                "zona_nombre": zona_nombre,
                "source_file": xlsx.name,
                "is_mock": "false",
            }
        )
    return clients, warnings


async def geocode_clients(
    clients: list[dict], api_key: str, barrio_hint: str
) -> None:
    async with httpx.AsyncClient() as client:
        for c in clients:
            if c["lat"] is not None and c["lng"] is not None:
                continue
            address = f"{c['domicilio']}, {barrio_hint}, Formosa, Formosa, Argentina"
            lat, lng, status = await geocode(client, address, api_key)
            if lat is None and api_key and not _GOOGLE_DISABLED:
                alt = f"{c['domicilio']}, Formosa, Formosa, Argentina"
                lat, lng, status = await geocode_google(client, alt, api_key)
                time.sleep(0.05)
            if lat is None:
                alt = f"{c['domicilio']}, Formosa, Argentina"
                lat, lng, status = await geocode_nominatim(client, alt)
                time.sleep(1.05)
            c["lat"], c["lng"] = lat, lng
            c["geocode_status"] = status
            time.sleep(0.08)


def write_zone_outputs(
    out: Path, zona: str, clients: list[dict], patch_zonas: bool
) -> None:
    slug = _slug(zona)
    csv_path = out / f"phase-04-clientes-diego-{slug}.csv"
    fieldnames = list(clients[0].keys()) if clients else []
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(clients)

    points = [
        (float(c["lat"]), float(c["lng"]))
        for c in clients
        if c["lat"] is not None
        and c["lng"] is not None
        and in_formosa(float(c["lat"]), float(c["lng"]))
    ]
    outliers = sum(
        1
        for c in clients
        if c["lat"] is not None
        and c["lng"] is not None
        and not in_formosa(float(c["lat"]), float(c["lng"]))
    )
    ok = sum(1 for c in clients if c["lat"] is not None)
    print(f"  [{zona}] clientes={len(clients)} coords={ok} outliers_bbox={outliers} → {csv_path.name}")

    if len(points) == 0:
        print(f"  [{zona}] hull omitido (sin puntos en bbox Formosa)")
        return
    # 1–2 puntos: buffer alrededor del punto / segmento (mismo buffer_m)
    gj = hull_geojson(points, buffer_m=300 if len(points) >= 3 else 400)
    geo_path = out / f"phase-04-zona-{slug}-polygon.geojson"
    wkt_path = out / f"phase-04-zona-{slug}-polygon.wkt"
    geo_path.write_text(json.dumps(gj, ensure_ascii=False, indent=2), encoding="utf-8")
    wkt = geojson_to_wkt(gj)
    wkt_path.write_text(wkt, encoding="utf-8")
    print(f"  [{zona}] hull → {geo_path.name}")

    if not patch_zonas:
        return
    zonas_csv = out / "phase-04-zonas-real.csv"
    if not zonas_csv.exists():
        return
    with zonas_csv.open(encoding="utf-8") as f:
        rows_z = list(csv.DictReader(f))
    dias = (clients[0].get("dias_zona") or "lunes").split(",")
    primary = dias[0]
    matched = False
    for z in rows_z:
        if z.get("nombre", "").strip().lower() == zona.strip().lower():
            z["geometry_wkt"] = wkt
            z["dias_visita"] = ",".join(dias)
            z["dia_visita"] = primary
            matched = True
    if matched:
        with zonas_csv.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(rows_z[0].keys()))
            w.writeheader()
            w.writerows(rows_z)
        print(f"  [{zona}] geometry patch en phase-04-zonas-real.csv")
    else:
        print(f"  [{zona}] aviso: no hallé zona en phase-04-zonas-real.csv")


async def main_async(
    esquema: str, dry_run: bool, only_zona: str | None, skip_existing_sf: bool
) -> None:
    base = ROOT / "implementacion" / esquema
    input_dir = base / "inputs" / "diego_ruta_1"
    out = base / "outputs"
    out.mkdir(parents=True, exist_ok=True)

    if not input_dir.is_dir():
        raise SystemExit(f"Falta {input_dir}")

    api_key = (os.getenv("GOOGLE_MAPS_API_KEY") or "").strip()
    files = discover_excels(input_dir)
    print(f"Excels Diego: {len(files)} en {input_dir}")

    all_clients: list[dict] = []
    all_warnings: list[str] = []

    for xlsx in files:
        zona, dias = ZONE_FROM_FILENAME[xlsx.name]
        if only_zona and zona.lower() != only_zona.strip().lower():
            continue
        clients, warnings = load_clients_from_xlsx(xlsx, zona, dias)
        all_warnings.extend(warnings)
        for w in warnings:
            print(f"  ! {w}")
        print(f"[*] {zona}: {len(clients)} filas ← {xlsx.name} dias={dias}")

        if dry_run:
            all_clients.extend(clients)
            continue

        if skip_existing_sf and zona.lower() == "san francisco":
            prev = out / "phase-04-clientes-diego-san-francisco.csv"
            if prev.exists():
                print(f"  reutilizo coords existentes {prev.name}")
                # merge lat/lng from previous by codigo
                prev_rows = {r["codigo"]: r for r in csv.DictReader(prev.open(encoding="utf-8"))}
                for c in clients:
                    p = prev_rows.get(c["codigo"])
                    if p and p.get("lat") and p.get("lng"):
                        c["lat"] = float(p["lat"])
                        c["lng"] = float(p["lng"])
                        c["geocode_status"] = p.get("geocode_status") or "resolved"

        need = sum(1 for c in clients if c["lat"] is None or c["lng"] is None)
        if need:
            if not api_key and not _GOOGLE_DISABLED:
                print("  aviso: sin GOOGLE_MAPS_API_KEY → Nominatim")
            print(f"  geocoding {need} direcciones…")
            await geocode_clients(clients, api_key, barrio_hint=zona)
        write_zone_outputs(out, zona, clients, patch_zonas=True)
        all_clients.extend(clients)

    # CSV consolidado (si --zona, mergea con CSVs por-zona ya existentes)
    if only_zona:
        merged: list[dict] = []
        for p in sorted(out.glob("phase-04-clientes-diego-*.csv")):
            if p.name == "phase-04-clientes-diego-all.csv":
                continue
            merged.extend(csv.DictReader(p.open(encoding="utf-8")))
        all_clients = merged
    if all_clients:
        cons = out / "phase-04-clientes-diego-all.csv"
        keys: list[str] = []
        for r in all_clients:
            for k in r:
                if k not in keys:
                    keys.append(k)
        with cons.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=keys)
            w.writeheader()
            w.writerows(all_clients)
        print(f"\nOK consolidado: {len(all_clients)} → {cons}")

    # resumen
    from collections import Counter

    print("\n## Resumen por zona")
    for zona, n in Counter(c["zona_nombre"] for c in all_clients).items():
        coords = sum(1 for c in all_clients if c["zona_nombre"] == zona and c.get("lat") not in (None, ""))
        print(f"  {zona}: {n} clientes" + ("" if dry_run else f" ({coords} con coords)"))
    if all_warnings:
        print(f"\nWarnings: {len(all_warnings)}")


def main() -> None:
    import asyncio

    ap = argparse.ArgumentParser()
    ap.add_argument("--esquema", default="el_gigante")
    ap.add_argument("--dry-run", action="store_true", help="Solo parsea; no geocode ni escribe hull")
    ap.add_argument("--zona", default=None, help="Procesar solo esta zona (ej. 'Centro 1')")
    ap.add_argument(
        "--skip-existing-sf",
        action="store_true",
        default=True,
        help="Reusar CSV San Francisco ya geocodificado (default on)",
    )
    ap.add_argument("--no-skip-existing-sf", action="store_true")
    args = ap.parse_args()
    skip_sf = not args.no_skip_existing_sf
    asyncio.run(main_async(args.esquema, args.dry_run, args.zona, skip_sf))


if __name__ == "__main__":
    main()
