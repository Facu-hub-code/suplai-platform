import os
import sys
import csv
import asyncio
import asyncpg
import requests
import openpyxl
from dotenv import load_dotenv

# Reconfigurar stdout a UTF-8 en Windows
if sys.platform.startswith('win'):
    sys.stdout.reconfigure(encoding='utf-8')

# Cargar variables de entorno
dotenv_path = r"c:\Users\marti\suplai-platform\.env"
load_dotenv(dotenv_path)

excel_path = r"c:\Users\marti\suplai-platform\implementacion\distribuidora_lyl\inputs\lista-productos.xlsx"
output_dir = r"c:\Users\marti\suplai-platform\implementacion\distribuidora_lyl\outputs"
schema = "distribuidora_lyl"
bucket_name = "products-distribuidora_lyl"

async def main():
    db_url = os.getenv("SUPABASE_DB_URL")
    supabase_url = os.getenv("SUPABASE_URL")
    supabase_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    
    if not db_url or not supabase_url or not supabase_key:
        print("[FAIL] Faltan variables de entorno requeridas (SUPABASE_DB_URL, SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY) en .env.")
        sys.exit(1)
        
    if not os.path.exists(excel_path):
        print(f"[FAIL] No existe el archivo de entrada {excel_path}")
        sys.exit(1)
        
    print(f"[*] Abriendo Excel: {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    # 1. Mapear imágenes en la hoja por fila (0-indexed)
    row_images = {}
    for img in ws._images:
        anchor = img.anchor
        row_idx = None
        if hasattr(anchor, '_from'):
            row_idx = anchor._from.row
        elif hasattr(anchor, 'row'):
            row_idx = anchor.row
            
        if row_idx is not None:
            row_images[row_idx] = img
            
    print(f"[*] Encontradas {len(ws._images)} imágenes en el archivo Excel.")
    
    # 2. Recorrer filas del Excel usando la misma lógica de procesar_excel_lyl.py
    sku_counter = 1
    product_image_mappings = {}  # product_code -> public_url
    product_codes = []
    
    print("[*] Emparejando productos con imágenes...")
    
    for row in range(1, ws.max_row + 1):
        val_col1 = ws.cell(row=row, column=1).value  # unidades_por_bulto (pack size)
        val_col3 = ws.cell(row=row, column=3).value  # nombre
        val_col5 = ws.cell(row=row, column=5).value  # precio unitario (price_unit)
        
        # Limpiar valor de nombre
        val_col3_str = str(val_col3).strip() if val_col3 is not None else ""
        
        # Saltear cabeceras de categorías
        if ">>>" in val_col3_str:
            continue
            
        # Saltear cabeceras e hilos vacíos
        if val_col3_str in ("", "nan", "Producto") or ("BEBIDAS" in val_col3_str and "ALCOHOL" in val_col3_str):
            continue
            
        # Validar precio
        try:
            price_unit = float(val_col5) if val_col5 is not None else 0
        except (ValueError, TypeError):
            continue
            
        if price_unit <= 0:
            continue
            
        # Fila identificada como producto
        product_code = f"LYL-{sku_counter:04d}"
        sku_counter += 1
        product_codes.append(product_code)
        
        # Comprobar imagen anclada en row-1 (0-indexed)
        anchor_row = row - 1
        if anchor_row in row_images:
            img_obj = row_images[anchor_row]
            
            # Obtener extensión
            fmt = (img_obj.format or "png").lower()
            if fmt == "jpeg":
                ext = "jpg"
            else:
                ext = fmt
                
            filename = f"{product_code}.{ext}"
            image_bytes = img_obj.ref.getvalue()
            
            # Subir a Supabase Storage
            upload_url = f"{supabase_url.rstrip('/')}/storage/v1/object/{bucket_name}/{filename}"
            headers = {
                "Authorization": f"Bearer {supabase_key}",
                "apikey": supabase_key,
                "Content-Type": f"image/{ext}",
                "x-upsert": "true"
            }
            
            try:
                resp = requests.post(upload_url, headers=headers, data=image_bytes)
                if resp.status_code == 200:
                    public_url = f"{supabase_url.rstrip('/')}/storage/v1/object/public/{bucket_name}/{filename}"
                    product_image_mappings[product_code] = public_url
                else:
                    print(f"[WARN] Error al subir imagen para {product_code} (HTTP {resp.status_code}): {resp.text}")
                    product_image_mappings[product_code] = "https://via.placeholder.com/150"
            except Exception as e:
                print(f"[WARN] Excepción al subir imagen para {product_code}: {e}")
                product_image_mappings[product_code] = "https://via.placeholder.com/150"
        else:
            # Producto sin imagen en el Excel
            product_image_mappings[product_code] = "https://via.placeholder.com/150"
            
    print(f"[*] Emparejamiento completo. Se subieron {len([url for url in product_image_mappings.values() if 'placeholder' not in url])} imágenes exitosamente.")
    
    # 3. Actualizar la base de datos Supabase
    print("[*] Conectando a Supabase DB para actualizar image_url...")
    conn = await asyncpg.connect(db_url)
    try:
        await conn.execute("SET statement_timeout = 0;")
        
        db_updates = []
        for code, img_url in product_image_mappings.items():
            db_updates.append((img_url, code))
            
        print(f"[*] Ejecutando actualización en batch para {len(db_updates)} productos...")
        await conn.executemany(f"""
            UPDATE {schema}.productos 
            SET image_url = $1, updated_at = now() 
            WHERE product_code = $2;
        """, db_updates)
        print("✅ Base de datos actualizada con las nuevas URLs de imagen.")
    except Exception as e:
        print(f"[FAIL] Error actualizando la base de datos: {e}")
        sys.exit(1)
    finally:
        await conn.close()
        
    # 4. Actualizar el archivo local CSV
    csv_path = os.path.join(output_dir, "phase-01-productos.csv")
    if os.path.exists(csv_path):
        print(f"[*] Actualizando archivo CSV local: {csv_path}...")
        
        # Leer líneas del CSV
        rows_to_write = []
        headers = []
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            for row_dict in reader:
                code = row_dict["product_code"]
                if code in product_image_mappings:
                    row_dict["image_url"] = product_image_mappings[code]
                rows_to_write.append(row_dict)
                
        # Escribir de vuelta al CSV
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows_to_write)
        print("✅ Archivo CSV local actualizado.")
    else:
        print(f"[WARN] No se encontró el CSV en {csv_path}, salteando actualización local.")
        
    # 5. Re-vectorización en el backend
    print("[*] Disparando la re-vectorización en el backend...")
    backend_url = os.getenv("BACKEND_URL", "https://web-production-f544f.up.railway.app").rstrip("/")
    vec_url = f"{backend_url}/{schema}/productos/vectorize"
    
    try:
        resp = requests.post(vec_url, json=product_codes, timeout=30)
        if resp.status_code == 200:
            print(f"✅ Re-vectorización encolada correctamente en el backend (HTTP 200).")
        else:
            print(f"[WARN] El backend devolvió HTTP {resp.status_code} al vectorizar: {resp.text}")
    except Exception as e:
        print(f"[WARN] Error al conectar con el endpoint de vectorización: {e}")
        
    print("\n==============================================")
    print("PROCESO DE ASOCIACIÓN DE IMÁGENES COMPLETADO")
    print("==============================================")

if __name__ == '__main__':
    asyncio.run(main())
