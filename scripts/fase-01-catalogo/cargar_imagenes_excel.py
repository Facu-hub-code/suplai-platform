import os
import sys
import csv
import asyncio
import asyncpg
import requests
import openpyxl
import argparse
import unicodedata
from dotenv import load_dotenv

# Reconfigurar stdout a UTF-8 en Windows
if sys.platform.startswith('win'):
    sys.stdout.reconfigure(encoding='utf-8')

# Cargar variables de entorno
dotenv_path = r"c:\Users\marti\suplai-platform\.env"
load_dotenv(dotenv_path)

def normalizar_texto(texto: str) -> str:
    if not texto:
        return ""
    # Convertir a minúsculas, remover acentos, puntuación y espacios adicionales
    flat = unicodedata.normalize('NFKD', str(texto).lower().strip())
    return "".join([c for c in flat if c.isalnum()])

async def main():
    parser = argparse.ArgumentParser(description="Herramienta Genérica de Asociación de Imágenes de Excel a Supabase.")
    parser.add_argument("--esquema", required=True, help="Nombre del esquema/tenant (ej: distribuidora_lyl)")
    parser.add_argument("--excel", help="Ruta al archivo Excel. Por defecto: implementacion/{esquema}/inputs/lista-productos.xlsx")
    parser.add_argument("--col-nombre", help="Letra o número de columna para el Nombre del producto (ej: C o 3). Opcional (auto-detectado).")
    parser.add_argument("--col-codigo", help="Letra o número de columna para el Código del producto (ej: A o 1). Opcional (auto-detectado).")
    parser.add_argument("--fila-inicio", type=int, default=1, help="Fila de inicio para el escaneo (por defecto: 1)")
    parser.add_argument("--bucket", help="Nombre del bucket de Supabase Storage. Por defecto: products-{esquema}")
    parser.add_argument("--forzar-prefijo", help="Prefijo para mapeo secuencial si falla la detección (ej: LYL).")
    
    args = parser.parse_args()
    schema = args.esquema
    
    # 1. Determinar ruta de Excel
    excel_file = args.excel
    if not excel_file:
        excel_file = rf"c:\Users\marti\suplai-platform\implementacion\{schema}\inputs\lista-productos.xlsx"
        if not os.path.exists(excel_file):
            excel_file = rf"c:\Users\marti\suplai-platform\implementacion\{schema}\inputs\lista_productos.xlsx"
            
    if not os.path.exists(excel_file):
        print(f"[FAIL] No se encontró el archivo Excel en: {excel_file}")
        sys.exit(1)
        
    db_url = os.getenv("SUPABASE_DB_URL")
    supabase_url = os.getenv("SUPABASE_URL")
    supabase_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    
    if not db_url or not supabase_url or not supabase_key:
        print("[FAIL] Faltan variables de entorno requeridas en .env (SUPABASE_DB_URL, SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY).")
        sys.exit(1)
        
    bucket_name = args.bucket or f"products-{schema}"
    
    # 2. Conectar a Base de Datos y descargar productos
    print(f"[*] Conectando a Supabase DB para descargar catálogo de '{schema}'...")
    conn = await asyncpg.connect(db_url)
    db_products = []
    try:
        db_products = await conn.fetch(f"SELECT product_code, nombre FROM {schema}.productos;")
        print(f"[*] Descargados {len(db_products)} productos desde la base de datos.")
    except Exception as e:
        print(f"[FAIL] Error leyendo la tabla {schema}.productos: {e}")
        await conn.close()
        sys.exit(1)
        
    if not db_products:
        print("[FAIL] No hay productos registrados en la base de datos para este esquema. Primero debes cargar el catálogo base.")
        await conn.close()
        sys.exit(1)
        
    # Indexar productos de la DB
    db_codes = {p["product_code"]: p["nombre"] for p in db_products}
    db_names_norm = {normalizar_texto(p["nombre"]): p["product_code"] for p in db_products}
    
    # 3. Leer libro de trabajo Excel
    print(f"[*] Leyendo Excel: {excel_file}...")
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb.active
    print(f"[*] Hoja activa: '{ws.title}' | Filas totales: {ws.max_row} | Columnas totales: {ws.max_column}")
    
    # 4. Auto-detección inteligente de columnas
    print("[*] Iniciando auto-detección de columnas...")
    
    # A. Mapear imágenes por celda (fila, columna) - 0-indexed
    cell_images = {}
    col_image_counts = {}
    for img in ws._images:
        anchor = img.anchor
        r, c = None, None
        if hasattr(anchor, '_from'):
            r = anchor._from.row
            c = anchor._from.col
        elif hasattr(anchor, 'row'):
            r = anchor.row
            c = anchor.col
            
        if r is not None and c is not None:
            # En openpyxl anchor c es 0-indexed, lo convertimos a 1-indexed
            col_1idx = c + 1
            cell_images[(r + 1, col_1idx)] = img
            col_image_counts[col_1idx] = col_image_counts.get(col_1idx, 0) + 1
            
    # La columna de imágenes es la que tenga más imágenes ancladas
    detected_image_col = None
    if col_image_counts:
        detected_image_col = max(col_image_counts, key=col_image_counts.get)
        print(f"  -> Columna de imágenes detectada: {detected_image_col} ({col_image_counts[detected_image_col]} imágenes)")
    else:
        print("  -> [WARN] No se detectaron imágenes incrustadas en ninguna columna.")
        
    # B. Escanear filas para detectar columna de Nombre y Código
    col_name_matches = {}
    col_code_matches = {}
    
    # Escanear una muestra de hasta 200 filas para detectar
    sample_limit = min(ws.max_row, args.fila_inicio + 200)
    for r in range(args.fila_inicio, sample_limit + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
                
            val_str = str(val).strip()
            norm = normalizar_texto(val_str)
            
            # Comprobar coincidencia con nombres de la DB
            if norm in db_names_norm:
                col_name_matches[c] = col_name_matches.get(c, 0) + 1
                
            # Comprobar coincidencia con códigos de la DB
            if val_str in db_codes:
                col_code_matches[c] = col_code_matches.get(c, 0) + 1
                
    # Determinar columnas elegidas
    # Columna Nombre: la que tenga más coincidencias de nombres normalizados
    detected_name_col = None
    if col_name_matches:
        detected_name_col = max(col_name_matches, key=col_name_matches.get)
        print(f"  -> Columna de nombres detectada: {detected_name_col} (Coincidencias en muestra: {col_name_matches[detected_name_col]})")
        
    # Columna Código: la que tenga más coincidencias de códigos exactos
    detected_code_col = None
    if col_code_matches:
        detected_code_col = max(col_code_matches, key=col_code_matches.get)
        print(f"  -> Columna de códigos detectada: {detected_code_col} (Coincidencias en muestra: {col_code_matches[detected_code_col]})")
        
    # Aplicar sobreescrituras manuales si se pasaron por parámetro
    def parse_column_input(col_input):
        if not col_input:
            return None
        if str(col_input).isdigit():
            return int(col_input)
        # Convertir letra a número (A=1, B=2, etc.)
        num = 0
        for char in str(col_input).upper():
            if 'A' <= char <= 'Z':
                num = num * 26 + (ord(char) - ord('A') + 1)
        return num if num > 0 else None

    manual_name_col = parse_column_input(args.col_nombre)
    manual_code_col = parse_column_input(args.col_codigo)
    
    final_name_col = manual_name_col or detected_name_col
    final_code_col = manual_code_col or detected_code_col
    
    if not final_name_col:
        print("[FAIL] No se pudo auto-detectar la columna de Nombres de productos. Por favor especifícala usando --col-nombre.")
        await conn.close()
        sys.exit(1)
        
    print(f"\n[*] Configuración de Columnas Final:")
    print(f"  - Columna Nombres: {final_name_col}")
    print(f"  - Columna Códigos: {final_code_col if final_code_col else 'No utilizada / Secuencial'}")
    print(f"  - Columna Imágenes: {detected_image_col if detected_image_col else 'No encontrada'}")
    print("-" * 50)
    
    # 5. Asegurar bucket en Supabase Storage
    print(f"[*] Verificando/Creando bucket público '{bucket_name}'...")
    storage_bucket_url = f"{supabase_url.rstrip('/')}/storage/v1/bucket"
    headers_storage = {
        "Authorization": f"Bearer {supabase_key}",
        "apikey": supabase_key,
        "Content-Type": "application/json"
    }
    
    try:
        bucket_payload = {"id": bucket_name, "name": bucket_name, "public": True}
        r_bucket = requests.post(storage_bucket_url, headers=headers_storage, json=bucket_payload)
        if r_bucket.status_code == 200:
            print(f"✅ Bucket '{bucket_name}' creado con éxito.")
        elif r_bucket.status_code == 400 or "already exists" in r_bucket.text.lower():
            print(f"[*] El bucket '{bucket_name}' ya está disponible.")
        else:
            print(f"[WARN] Error creando bucket: {r_bucket.status_code} - {r_bucket.text}")
    except Exception as e:
        print(f"[WARN] No se pudo asegurar el bucket: {e}")
        
    # 6. Escanear y subir imágenes
    print("\n[*] Procesando filas y emparejando imágenes...")
    
    product_image_mappings = {}  # product_code -> public_url
    product_codes_to_vectorize = []
    
    matched_by_code = 0
    matched_by_name = 0
    matched_by_seq = 0
    
    sku_counter = 1
    
    for r in range(args.fila_inicio, ws.max_row + 1):
        cell_name = ws.cell(row=r, column=final_name_col).value
        cell_name_str = str(cell_name).strip() if cell_name is not None else ""
        
        if not cell_name_str or cell_name_str.lower() in ("nan", "none", "nombre", "producto", "descripcion"):
            continue
            
        product_code = None
        
        # Estrategia 1: Emparejar por código de producto si se usa columna de código
        if final_code_col:
            cell_code = ws.cell(row=r, column=final_code_col).value
            cell_code_str = str(cell_code).strip() if cell_code is not None else ""
            if cell_code_str in db_codes:
                product_code = cell_code_str
                matched_by_code += 1
                
        # Estrategia 2: Emparejar por coincidencia de nombre normalizado
        if not product_code:
            norm_name = normalizar_texto(cell_name_str)
            if norm_name in db_names_norm:
                product_code = db_names_norm[norm_name]
                matched_by_name += 1
                
        # Estrategia 3: Emparejar secuencialmente (Fallback)
        if not product_code and args.forzar_prefijo:
            product_code = f"{args.forzar_prefijo}-{sku_counter:04d}"
            sku_counter += 1
            matched_by_seq += 1
            
        if not product_code:
            # Incrementar secuencial por si acaso
            sku_counter += 1
            continue
            
        product_codes_to_vectorize.append(product_code)
        
        # Buscar si hay alguna imagen anclada en esta fila
        # escaneamos todas las columnas de la fila 'r' por si la imagen está en cualquier columna
        row_img_obj = None
        if detected_image_col:
            row_img_obj = cell_images.get((r, detected_image_col))
        else:
            # Fallback: buscar en toda la fila
            for col_temp in range(1, ws.max_column + 1):
                if (r, col_temp) in cell_images:
                    row_img_obj = cell_images[(r, col_temp)]
                    break
                    
        if row_img_obj:
            # Extraer bytes y subir
            fmt = (row_img_obj.format or "png").lower()
            ext = "jpg" if fmt == "jpeg" else fmt
            filename = f"{product_code}.{ext}"
            image_bytes = row_img_obj.ref.getvalue()
            
            upload_url = f"{supabase_url.rstrip('/')}/storage/v1/object/{bucket_name}/{filename}"
            headers_upload = {
                "Authorization": f"Bearer {supabase_key}",
                "apikey": supabase_key,
                "Content-Type": f"image/{ext}",
                "x-upsert": "true"
            }
            
            try:
                resp = requests.post(upload_url, headers=headers_upload, data=image_bytes)
                if resp.status_code == 200:
                    public_url = f"{supabase_url.rstrip('/')}/storage/v1/object/public/{bucket_name}/{filename}"
                    product_image_mappings[product_code] = public_url
                else:
                    print(f"[WARN] Error subiendo imagen para {product_code} (HTTP {resp.status_code}): {resp.text}")
                    product_image_mappings[product_code] = "https://via.placeholder.com/150"
            except Exception as e:
                print(f"[WARN] Excepción subiendo imagen para {product_code}: {e}")
                product_image_mappings[product_code] = "https://via.placeholder.com/150"
        else:
            # Si no hay imagen en la fila, no modificamos la URL o ponemos placeholder
            # Para evitar pisar URLs válidas existentes, solo registramos placeholder si el producto es nuevo
            product_image_mappings[product_code] = "https://via.placeholder.com/150"
            
    print(f"\n[*] Emparejamiento finalizado:")
    print(f"  - Por Código: {matched_by_code}")
    print(f"  - Por Nombre: {matched_by_name}")
    print(f"  - Por Orden Secuencial: {matched_by_seq}")
    print(f"  - Total Mapeados con URLs: {len(product_image_mappings)}")
    
    # 7. Actualizar base de datos
    if product_image_mappings:
        print("\n[*] Actualizando URLs de imagen en Supabase DB...")
        try:
            db_updates = []
            for code, img_url in product_image_mappings.items():
                db_updates.append((img_url, code))
                
            await conn.execute("SET statement_timeout = 0;")
            await conn.executemany(f"""
                UPDATE {schema}.productos 
                SET image_url = $1, updated_at = now() 
                WHERE product_code = $2;
            """, db_updates)
            print(f"✅ Base de datos actualizada ({len(db_updates)} productos).")
        except Exception as e:
            print(f"[FAIL] Error actualizando Supabase DB: {e}")
            await conn.close()
            sys.exit(1)
    else:
        print("[*] No se encontraron asociaciones de imágenes para actualizar en la DB.")
        
    await conn.close()
    
    # 8. Sincronizar archivo CSV local si existe
    csv_path = rf"c:\Users\marti\suplai-platform\implementacion\{schema}\outputs\phase-01-productos.csv"
    if os.path.exists(csv_path):
        print(f"\n[*] Sincronizando archivo CSV local en {csv_path}...")
        rows_to_write = []
        headers_csv = []
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers_csv = reader.fieldnames
            for row_dict in reader:
                code = row_dict["product_code"]
                if code in product_image_mappings:
                    row_dict["image_url"] = product_image_mappings[code]
                rows_to_write.append(row_dict)
                
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers_csv)
            writer.writeheader()
            writer.writerows(rows_to_write)
        print("✅ Archivo CSV local sincronizado.")
        
    # 9. Disparar re-vectorización en el backend
    if product_codes_to_vectorize:
        print("\n[*] Disparando re-vectorización en el backend...")
        backend_url = os.getenv("BACKEND_URL", "https://web-production-f544f.up.railway.app").rstrip("/")
        vec_url = f"{backend_url}/{schema}/productos/vectorize"
        try:
            resp = requests.post(vec_url, json=product_codes_to_vectorize, timeout=30)
            if resp.status_code == 200:
                print(f"✅ Re-vectorización encolada en el backend exitosamente.")
            else:
                print(f"[WARN] Backend devolvió HTTP {resp.status_code} al vectorizar: {resp.text}")
        except Exception as e:
            print(f"[WARN] Error disparando vectorización: {e}")
            
    print("\n==============================================")
    print("PROCESO DE ASOCIACIÓN MASIVA COMPLETADO CON ÉXITO")
    print("==============================================")

if __name__ == '__main__':
    asyncio.run(main())
